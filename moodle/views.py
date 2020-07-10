import base64
from openpyxl import Workbook, load_workbook
import hmac
import time
import uuid
import jwt
import datetime

from django.core.files.base import ContentFile
from django.db.models import Q, F
from django.http import HttpResponse,JsonResponse
from django.shortcuts import redirect, render
from django.views import View
# import smtplib
# from email.mime.text import MIMEText
# from email.header import Header
from collections import Counter

from moodle.models import *
from users.models import Users
from tool.session_token import check_session_token, make_session_token

# ▼ ▼ ▼ ▼ ▼ ▼ ▼ ▼ ▼ ▼ ▼ ▼ ▼ ▼ 基础设置 ▼ ▼ ▼ ▼ ▼ ▼ ▼ ▼ ▼ ▼ ▼ ▼ ▼ ▼


salt = b'cso_canada_summer'  # 密码盐

salt_t = 'cso_canada_summer_w'  #token盐
expire = 6  #token超时时间


class Role:
    '''
    权限等级
    现有权限等级(将新增 role 字段放入对应等级列表中):
        普通权限 PermissionsL : 仅可获取自己(token中对应的)obj对象数据
        I级权限 PermissionsI : 可获取任意指定对象数据
    '''

    def __init__(self):
        self.PermissionsL = ['stu']
        self.PermissionsI = ['prof']
        # self.PermissionsII = []
        # self.PermissionsIII = []


# ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ 基础设置 ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲

pass



# ▼ --------------------------------------- ▼  user_views  ▼ --------------------------------------- ▼


# class MoodleUsersMsg(View):
#     '''
#     所有用户/注册
#     '''
#
#     @check_token
#
#     def post(self, request):
#         '''
#         注册
#         :param request:
#         :return:
#         '''
#
#         # json格式检查
#         try:
#             account = request.data['m_ac']
#             password = request.data['m_ps']
#             E_code = request.data['E_code']
#         except:
#             return JsonResponse({
#                 'code': 40407,
#                 'data': 'json参数有误'
#             })
#
#         # 查重
#         try:
#             MoodleUser.objects.get(m_ac=account)
#             return JsonResponse({
#                 'code': 40401,
#                 'data': '该账户已存在,无法注册'
#             })
#         except:
#             pass
#
#         # 验证码验证
#         try:
#             E_obj = EmailCode.objects.get(Email=account)
#             E_time = E_obj.E_time
#             interval = float(time.time() - float(E_time))
#             # 超时验证 10min
#             if interval > 600:
#                 return JsonResponse({
#                     'code': 40409,
#                     'data': '验证码已失效'
#                 })
#
#             # 验证码相同 进行注册
#             if E_code == E_obj.E_code:
#                 h = hmac.new(salt, password.encode(), digestmod='sha256')
#                 m_ps_h = h.hexdigest()
#
#                 new_data = {
#                     'm_ac': account,
#                     'm_ps': m_ps_h
#                 }
#
#                 serializer = UserSerializer(data=new_data)
#                 if serializer.is_valid():
#                     q_obj = serializer.save()
#
#                     # 生成token
#                     token = make_token(m_ac_id=q_obj.pk)
#
#                     return JsonResponse({
#                         'code': 200,
#                         'data': {'token': token.decode()}
#                     })
#                 else:
#                     return JsonResponse({
#                         'code': 40402,
#                         'data': serializer.errors
#                     })
#
#             # 验证码不相同
#             else:
#                 return JsonResponse({
#                     'code': 40410,
#                     'data': '验证码错误'
#                 })
#
#         except:
#             # 数据库无当前用户验证码数据
#             return JsonResponse({
#                 'code': 40408,
#                 'data': '未获取验证码'
#             })






# class MoodleUserMsg(View):
#     '''
#     指定用户
#     '''
#
#     @check_token
#     def get(self, request, pk=None):
#         role = request.role
#
#         # 普通权限或无查询字串
#         if (role in Role().PermissionsL) or (pk == None):
#             serializer = UserSerializer(instance=request.user_obj)
#             return JsonResponse({
#                 'code': 200,
#                 'data': serializer.data
#             })
#
#         # I级权限
#         elif role in Role().PermissionsI:
#             try:
#                 q_obj = MoodleUser.objects.get(id=pk)
#             except:
#                 return JsonResponse({
#                     'code': 40412,
#                     'data': '你要查询的用户不存在'
#                 })
#             serializer = UserSerializer(instance=q_obj)
#             return JsonResponse({
#                 'code': 200,
#                 'data': serializer.data
#             })
#
#     @check_token
#     def put(self, request, pk=None):
#         role = request.role
#
#         # 普通权限或无查询字串
#         if (role in Role().PermissionsL) or (pk == None):
#             serializer = UserSerializer(instance=request.user_obj, data=request.data)
#             if serializer.is_valid():
#                 _q_obj = serializer.save()
#                 return JsonResponse({
#                     'code': 200,
#                     'data': _q_obj.pk
#                 })
#             else:
#                 return JsonResponse({
#                     'code': 40404,
#                     'data': serializer.errors
#                 })
#
#         # I级权限
#         elif role in Role().PermissionsI:
#             try:
#                 q_obj = MoodleUser.objects.get(id=pk)
#             except:
#                 return JsonResponse({
#                     'code': 40403,
#                     'data': '你要修改的用户不存在'
#                 })
#
#             serializer = UserSerializer(instance=q_obj, data=request.data)
#             if serializer.is_valid():
#                 _q_obj = serializer.save()
#                 return JsonResponse({
#                     'code': 200,
#                     'data': _q_obj.pk
#                 })
#             else:
#                 return JsonResponse({
#                     'code': 40404,
#                     'data': serializer.errors
#                 })
#
#     # def delete(self,request,m_ac):
#     #     try:
#     #         q_obj = Moodle.objects.get(m_account=m_ac)
#     #     except:
#     #         return JsonResponse({
#     #             'code': 40405,
#     #             'data': '你要删除的用户不存在'
#     #         })
#     #
#     #     q_obj.delete()
#     #     return JsonResponse({
#     #         'code':200,
#     #         'data':'删除成功'
#     #     })
#
# # ▲ --------------------------------------- ▲  user_views  ▲ --------------------------------------- ▲
# pass
#
#
# # ▼ --------------------------------------- ▼ course_views ▼ --------------------------------------- ▼
#
# class MoodleLearnData(View):
#     '''
#     用户学习数据
#     '''
#
#     @check_token
#     def put(self,request,judge):
#
#         #获取CUS对象
#         try:
#             course = int(request.data['course'])
#             json_data = request.data['data']
#             CUS_obj = request.user_obj.m_cus.get(moodle_course=course)
#         except:
#             return JsonResponse({
#                 'code':40417,
#                 'data':'json参数有误'
#             })
#
#         #修改用户观看视频时长
#         if judge == 'time':
#             # 重写data
#             time = int(json_data) + CUS_obj.time
#             data = {
#                 'time': time
#             }
#
#             serializer = CUSSerializer(instance=CUS_obj,data=data)
#             if serializer.is_valid():
#                 serializer.save()
#                 return JsonResponse({
#                     'code': 200,
#                     'data': 'OK'
#                 })
#             else:
#                 return JsonResponse({
#                     'code': 40416,
#                     'data': serializer.errors
#                 })
#
#         #修改用户观看视频进度
#         elif judge == 'progress':
#
#             video_num = len(CUS_obj.moodle_course.m_video.all())
#             data_progress = int(json_data) / video_num
#             if data_progress > 1:
#                 return JsonResponse({
#                     'code':40418,
#                     'data':'data值不在范围内'
#                 })
#             print(video_num , data_progress)
#
#             if data_progress <= CUS_obj.progress:
#                 return JsonResponse({
#                     'code':200,
#                     'data':'PASS'
#                 })
#
#             else:
#                 # 重写data
#                 data = {
#                     'progress': data_progress
#                 }
#
#                 serializer = CUSSerializer(instance=CUS_obj, data=data)
#                 if serializer.is_valid():
#                     serializer.save()
#                     return JsonResponse({
#                         'code': 200,
#                         'data': 'OK'
#                     })
#                 else:
#                     return JsonResponse({
#                         'code': 40416,
#                         'data': serializer.errors
#                     })
#
#         #API不存在
#         else:
#             return JsonResponse({
#                 'code':404,
#                 'data':'API不存在'
#             })
#
#
# class MoodleCoursesMsg(View):
#     '''
#     所有课程(学科)
#     '''
#
#     @check_token
#     def get(self, request):
#
#         #获取学科查询结果集
#         s_q_set = MoodleSubject.objects.all()
#         #新建学科-课程列表
#         sub_cou_list = []
#
#         #遍历学科查询结果集
#         for s in s_q_set:
#             #序列化学科数据
#             s_dic = SubjectSerializer(instance=s).data
#             #序列化学科旗下课程数据
#             c_dic = CourseSerializer(instance=s.m_course.all(),many=True,excludes=['subject','m_user']).data
#             #学科数据中添加课程数据
#             s_dic['course'] = c_dic
#             #学科-课程列表中添加更改后的学科数据
#             sub_cou_list.append(s_dic)
#
#
#
#         return JsonResponse({
#             'code':200,
#             'data':sub_cou_list
#         })
#
#
# class MoodleCourseMsg(View):
#     '''
#     我的课程/指定学员的课程
#     '''
#
#     @check_token
#     def get(self,request,pk=None):
#         if not pk:
#             user_obj = request.user_obj
#             my_course = user_obj.m_course.all()
#
#             serializer = CourseSerializer(instance=my_course,many=True)
#             for i in serializer.data:
#                 cus_obj = user_obj.m_cus.get(moodle_course = i['id'])
#
#                 i['time'] = cus_obj.time
#                 i['progress'] = cus_obj.progress
#
#             return JsonResponse({
#                 'code':200,
#                 'data':serializer.data
#             })
#
#         else:
#             try:
#                 user_obj = MoodleUser.objects.get(id=pk)
#             except:
#                 return JsonResponse({
#                     'code': 40412,
#                     'data': '你要查询的用户不存在'
#                 })
#             my_course = user_obj.m_course.all()
#             serializer = CourseSerializer(instance=my_course,many=True)
#             return JsonResponse({
#                 'code': 200,
#                 'data': serializer.data
#             })
#
# # ▲ --------------------------------------- ▲ course_views ▲ --------------------------------------- ▲


class Login(View):
    '''
    用户登录
    '''
    # def post(self, request):
    #
    #     # json格式检查
    #     try:
    #         password = request.data['m_ps']
    #         account = request.data['m_ac']
    #     except:
    #         return JsonResponse({
    #             'code': 40413,
    #             'data': 'json参数有误'
    #         })
    #
    #     # 数据库查询
    #     try:
    #         user_obj = MoodleUser.objects.get(m_ac=account)
    #         h = hmac.new(salt, password.encode(), digestmod='sha256')
    #         m_ps_h = h.hexdigest()
    #
    #         # 密码比对
    #         if user_obj.m_ps == m_ps_h:
    #             token = make_token(m_ac_id=user_obj.pk)
    #             return JsonResponse({
    #                 'code': 200,
    #                 'data': {'token': token.decode()}
    #             })
    #
    #         else:
    #             return JsonResponse({
    #                 'code': 40414,
    #                 'data': '密码错误'
    #             })
    #
    #     except Exception as e:
    #         print(e)
    #         return JsonResponse({
    #             'code': 40415,
    #             'data': '你要登陆的用户不存在'
    #         })

    def get(self,request):

        return render(request,'moodle/login.html')

    def post(self, request):

        # 格式检查
        try:

            account = request.POST['m_ac']
            password = request.POST['m_ps']
            h = hmac.new(salt, password.encode(), digestmod='sha256')
            m_ps_h = h.hexdigest()
        except:
            return JsonResponse({
                'code': 40413,
                'data': 'json参数有误'
            })

        # 数据库查询
        # 判断小程序数据库有无数据
        try:
            # 小程序有数据
            user_obj = Users.objects.get(firstEmail=account,isActive='true')

            # 判断moodle数据库有无数据
            try:
                MoodleUser.objects.get(m_ac=account)
            except:

                # 不存在创建数据
                new_m_user = MoodleUser()
                new_m_user.m_ac = user_obj.firstEmail
                h = hmac.new(salt, user_obj.domesticTelephone.encode(), digestmod='sha256')
                new_m_ps_h = h.hexdigest()
                new_m_user.m_ps = new_m_ps_h
                new_m_user.users = user_obj
                new_m_user.nick = user_obj.username
                new_m_user.save()

            # 判断账号密码是否正确
            m_user_obj = MoodleUser.objects.get(m_ac=account)

            if m_user_obj.m_ps == m_ps_h:
                # 正确
                token = make_session_token(m_ac_id=m_user_obj.id)

                return_obj = redirect("course")

                return_obj.set_cookie('token',token.decode())

                # request.session['authorization'] = {'token': token.decode(), 'user': m_user_obj.nick, 'role': m_user_obj.role}

                return return_obj

            else:

                 return render(request,'moodle/login.html', {'msg': "密码错误"})

        except:
            # 小程序无数据 判断moodle数据库有无数据

            try:
                # moodle数据库有数据
                MoodleUser.objects.get(m_ac=account)
                # 判断账号密码是否正确
                m_user_obj = MoodleUser.objects.get(m_ac=account)

                if m_user_obj.m_ps == m_ps_h:
                    # 正确
                    token = make_session_token(m_ac_id=m_user_obj.id)

                    return_obj = redirect("course")

                    return_obj.set_cookie('token', token.decode())

                    return return_obj


                else:
                    # moodle数据库无数据
                     return render(request,'moodle/login.html', {'msg': "密码错误"})
            except:
                pass

            return render(request,'moodle/login.html', {'msg': "密码错误"})



# class Index(View):
#
#     def get(self,request):
#         return render(request,'moodle/index.html')



class Contact(View):

    @check_session_token
    def get(self,request):
        return render(request,'moodle/contact.html', {'notice': MoodleNotice.objects.all()})


class User(View):

    @check_session_token
    def get(self, request):
        # print(request.user_obj.__dict__)

        notice = MoodleNotice.objects.all()

        user = request.user_obj.users
        try:
            school = request.user_obj.users.school.get(func='zs')
        except:
            school = '暂无信息'

        return render(request, 'moodle/user.html', locals())


# -----------------#
class Correct(View):
    '''
    教授批改作业，学生答卷，学生查看教授批阅的答卷
    '''
    @check_session_token
    def get(self, request, exam_name, course_id):

        user_obj = request.user_obj

        # 确认课程权限
        queryset = user_obj.m_cus.filter(moodle_course__id=course_id)

        exam_info = MoodleExam.objects.filter(Q(exam_name=exam_name) & Q(moodle_course__id=course_id)).first()

        if queryset:

            if user_obj.role == "prof":

                course = user_obj.m_course.filter(id=course_id).first()
                all_stu_list = course.m_cus.filter(moodle_user__role='stu')


                exam_list = MoodleExamStu.objects.filter(Q(exam_name=exam_name) & Q(moodle_course=course_id) & ~Q(exam_answer='',content=None))

                exam_user_list = [user.moodle_user for user in exam_list]

                no_sub = [all_user for all_user in all_stu_list if all_user.moodle_user not in exam_user_list]

                this_time = datetime.datetime.now()

                if (this_time > exam_info.end_time):
                    show = 'show'
                else:
                    show = ''


                return render(request, 'moodle/correct.html', {"exam_info":exam_info, "exam_list": exam_list, 'notice': MoodleNotice.objects.all(),'no_sub':no_sub,'show':show})

            else:

                linshi = user_obj.m_linshi_stu.filter(Q(exam_name=exam_name) & Q(moodle_course__id=course_id)).first()

                stu_exam = user_obj.m_exam_stu.filter(Q(exam_name=exam_name) & Q(moodle_course__id=course_id)).first()

                if stu_exam:
                    tt = stu_exam.exam_answer or stu_exam.content
                else:
                    tt = ""

                # print("exam", stu_exam)


                # 当前时间
                this_time = datetime.datetime.now()

                if (this_time > exam_info.end_time) or tt:
                    # 当前时间大于考试结束时间。考试结束

                    exam_state = "closed"
                    return render(request, 'moodle/my_exam.html', {"stu_exam": stu_exam, 'user': user_obj, 'notice': MoodleNotice.objects.all()})

                elif this_time > exam_info.start_time:
                    # 当前时间大于考试时间,考试开放
                    exam_state = "open"

                else:
                    # 考试未开放
                    exam_state = "Unopen"

                return render(request, 'moodle/exam.html', {"this_time": this_time, "exam_state": exam_state, "exam_info": exam_info, 'user': user_obj, 'notice': MoodleNotice.objects.all(), 'linshi': linshi})


    @check_session_token
    def post(self, request, exam_name, course_id):

        user_obj = request.user_obj

        ExamStu = MoodleExamStu.objects.filter(exam_name=exam_name,moodle_user=user_obj,moodle_course__id=course_id)

        if ExamStu:
            ExamStu = ExamStu.first()
        else:
            ExamStu = MoodleExamStu()

        ExamStu.exam_name = exam_name

        ExamStu.moodle_user = user_obj

        ExamStu.moodle_course = MoodleCourse.objects.get(id=course_id)



        try:

            image = request.POST['image']

            e_type = request.POST['type']
            if e_type == 'online':
                # result = image.split(',')

                # image_data = base64.b64decode(result[1])
                # imagene = ContentFile(image_data, exam_name + str(user_obj.id) + '.jpg')

                # ExamStu.exam_answer = imagene

                ExamStu.content = image.replace("\\", "╲")

                ExamStu.save()

                return JsonResponse({"save": True})


        except:

            try:

                # print(request.FILES['image'])

                image = request.FILES['image']

                ExamStu.exam_answer = image

                ExamStu.save()

                return redirect(request.path_info)
            except Exception as e:
                print('--------------------',e)



class Marking(View):
    '''
    阅卷
    '''
    @check_session_token
    def get(self, request, course_id, exam_id):

        this_time = datetime.datetime.now()

        # print(request.GET['stu_id'])

        user_obj = request.user_obj

        if exam_id == 'None':
            queryset = user_obj.m_cus.filter(moodle_course__id=course_id)


            try:
                moodle_user_obj = MoodleUser.objects.get(id=request.GET['stu_id'])



                stu_exam = MoodleExamStu.objects.get(moodle_user=moodle_user_obj,exam_name=request.GET['exam_name'],moodle_course=queryset.first().moodle_course)



            except Exception as e:
                stu_exam = MoodleExamStu()


            if queryset:
                if user_obj.role == "prof":

                    

                    # exam_info = user_obj.m_exam.filter(exam_name=request.GET['exam_name']).first()
                    exam_info = MoodleExam.objects.filter(Q(exam_name=request.GET['exam_name']) & Q(moodle_course__id=course_id)).first()

                    stu_exam.exam_name = request.GET['exam_name']
                    stu_exam.moodle_user = MoodleUser.objects.get(id=request.GET['stu_id'])
                    stu_exam.moodle_course = queryset.first().moodle_course

                    stu_exam.save()

                    try:
                        return render(request, 'moodle/exam.html', {'this_time':this_time, 'stu_exam': stu_exam, 'user': user_obj, 'notice': MoodleNotice.objects.all(), 'exam_info':exam_info})
                    except Exception as e:
                        return HttpResponse(e)


        else:

            # 确认课程权限
            queryset = user_obj.m_cus.filter(moodle_course__id=course_id)
            stu_exam = MoodleExamStu.objects.filter(id=exam_id).first()

            if queryset:
                if user_obj.role == "prof":
                    
                    # exam_info = user_obj.m_exam.filter(exam_name=request.GET['exam_name']).first()
                    exam_info = MoodleExam.objects.filter(Q(exam_name=request.GET['exam_name']) & Q(moodle_course__id=course_id)).first()


                    return render(request, 'moodle/exam.html', {'this_time':this_time, 'stu_exam': stu_exam, 'user': user_obj, 'notice': MoodleNotice.objects.all(), 'exam_info':exam_info})

    @check_session_token
    def post(self, request, course_id, exam_id):
        # print(1111111111111111111111111111111111111,exam_id,request.path)
        queryset = MoodleExamStu.objects.filter(id=exam_id).first()

        score = request.POST["score"]
        if not score:
            score = 0

        if queryset:
            try:
                e_type = request.POST['type']
                if e_type == 'online':
                    image = request.POST['image']

                    result = image.split(',')

                    image_data = base64.b64decode(result[1])
                    imagene = ContentFile(image_data, exam_id+'.jpg')

                    queryset.exam_results = imagene
                    queryset.exam_score = float(score)
                    queryset.save()
                    return JsonResponse({"save": True})
            except:

                try:
                    image = request.FILES['image']
                    queryset.exam_results = image
                except:
                    queryset.exam_results = 'py.jpg'
                    queryset.exam_answer = 'py.jpg'

                queryset.exam_score = float(score)
                queryset.save()
                
                url_path =f"/v1/moodle/correct/{course_id}/{queryset.exam_name}"

                return redirect(url_path)


class Credits(View):
    '''
    学分认证
    '''

    @check_session_token
    def get(self, request):
        user_obj = request.user_obj

        grade_list = user_obj.m_cus.all()

        return render(request, 'moodle/credits.html', {'grade_list': grade_list, 'user': user_obj, 'notice': MoodleNotice.objects.all()})


class Score(View):
    '''
    教授设置学生总成绩
    '''

    @check_session_token
    def get(self, request):
        user_obj = request.user_obj
        course_list = user_obj.m_course.all()

        return render(request, 'moodle/score.html', {'course_list': course_list, 'notice': MoodleNotice.objects.all()})


class ShowScore(View):
    '''
    教授查看某课所有学生列表
    '''
    @check_session_token
    def get(self, request, course_id):
        user_obj = request.user_obj
        course = user_obj.m_course.filter(id=course_id).first()
        all_stu_list = course.m_cus.filter(moodle_user__role='stu')

        try:

            D = []
            for stu in all_stu_list:
                st = {}
                st['stu'] = stu
                st['exam_list'] = []


                sum_list = []

                for exam in stu.moodle_user.m_exam_stu.filter(moodle_course__id=course_id):
                    exam_list = {}
                    exam_list['exam'] = exam

                    score = exam_list['exam'].exam_score
                    if not score:
                        score = 0


                    queryset = MoodleExam.objects.filter(Q(exam_name=exam_list['exam'].exam_name) & Q(moodle_course__id=course_id)).first()
                    exam_list['rate'] = queryset.rate
                    exam_list['result'] = float(score)*int(exam_list['rate'])/100

                    st["exam_list"].append(exam_list)

                    sum_list.append(exam_list['result'])


                st['Cal_grade'] = sum(sum_list)

                D.append(st)

            return render(request, 'moodle/setscore.html', {'all_stu_list': all_stu_list, 'course': course, 'notice': MoodleNotice.objects.all(), 'D':D})
        except Exception as e:
            print('-------报错-------------',e)

        # return render(request, 'moodle/setscore.html', {'all_stu_list': all_stu_list, 'course': course, 'notice': MoodleNotice.objects.all()})


    @check_session_token
    def post(self, request, course_id):

        stu_id = request.POST["stu_id"]

        score = request.POST["score"]

        queryset = MoodleCUS.objects.filter(Q(moodle_course=course_id) & Q(moodle_user=stu_id)).first()

        try:
            if queryset:
                queryset.grade = float(score)
                queryset.save()

            return JsonResponse({"save": True})
        
        except Exception as e:
            print("-----post---", e)


class Course(View):
    '''
    我的课程
    '''
    @check_session_token
    def get(self, request):
        user_obj = request.user_obj

        course_list = user_obj.m_course.all()

        return render(request, 'moodle/course.html', {"course_list": course_list, 'notice': MoodleNotice.objects.all()})


class Index(View):
    '''
    首页
    '''
    def get(self, request):

        try:
            token = request.COOKIES['token']

            print('--------------------',token)
            res = jwt.decode(token.encode(), salt_t, algorithms='HS256')
            m_ac_id = res['m_ac_id']
            m_user_obj = MoodleUser.objects.get(id=m_ac_id)
            # m_user_obj = user_obj
            # print(m_user_obj)
            setattr(request, 'user_obj', m_user_obj)
            setattr(request, 'role', m_user_obj.role)
        except Exception as e:
            print('111111111111111111111111111111111',e)

        hot_course = MoodleCourse.objects.filter(is_hot='T')[:10]
        hot_course_5 = hot_course[:5]
        hot_course_10 = hot_course[5:10]
        active_list = MoodleActive.objects.all()

        return render(request, 'moodle/index.html', {"hot_course_5": hot_course_5, "hot_course_10": hot_course_10, "active_list": active_list, 'notice': MoodleNotice.objects.all()})


class Exam(View):
    @check_session_token
    def post(self,request, exam_name, course_id):

        user_obj = request.user_obj

        linshi = user_obj.m_linshi_stu.filter(Q(exam_name=exam_name) & Q(moodle_course__id=course_id)).first()

        if linshi:
            linshi = linshi

        else:
            linshi = Linshi()

        # print(request.POST['lsbc'])
        try:

            linshi.moodle_user = user_obj
            linshi.exam_name = exam_name
            linshi.moodle_course = MoodleCourse.objects.get(id=course_id)

            linshi.content = request.POST['lsbc'].replace("\\", "╲")
            linshi.number = request.POST['bj_num']
            linshi.save()
        except:
            pass

        return JsonResponse({"save": True})


class Player(View):
    @check_session_token
    def get(self,request,course_id):

        try:

            user_obj = request.user_obj

            course = user_obj.m_course.filter(id=course_id)

            # 验证是否选修这门课
            if course:
                course_info = course.first()

                # print(course,course_info)

                video_list = course_info.m_video.all().order_by('v_index', 'v_name')

                # print(course_info.id)

                file_list = course_info.m_courseware.all()

                # 已结束
                exam_list_end = course_info.m_exam.filter(Q(moodle_user__role='prof') & Q(end_time__lt=datetime.datetime.now()))
                # 未开始
                exam_list_start = course_info.m_exam.filter(Q(moodle_user__role='prof') & Q(start_time__gte=datetime.datetime.now()))
                # 正在开始
                exam_list_open = course_info.m_exam.filter(Q(moodle_user__role='prof') & Q(start_time__lte=datetime.datetime.now()) & Q(end_time__gte=datetime.datetime.now()))

                # 已考过
                exam_list_kaowan = list(user_obj.m_exam_stu.filter(moodle_course__id=course_id))

                # print(11111111111111111111111111)

            else:
                video_list = ''
                file_list = ''
                course_info = ''
                exam_list_end = ''
                exam_list_start = ''
                exam_list_open = ''
                exam_list_kaowan = ''

            # print(video_list,file_list,exam_list_end)

            # if user_obj.active_date == datetime.datetime.strptime('3000-01-01 00:00:00', '%Y-%m-%d %H:%M:%S'):
            #     print ('OKOKOKOKOKOKOKOKOKOKOKOKOKOKOKOKOKOKOKOKOKK')
            # else:
            #     print ('!!!!!!!!!!!!!!!!!',type(user_obj.active_date))


            if (user_obj.active_date == datetime.datetime.strptime('3000-01-01 00:00:00', '%Y-%m-%d %H:%M:%S')) and (user_obj.role != 'prof'):
                return render(request, 'moodle/player.html', {'user': user_obj, "video_list": '', "file_list": '', "exam_list_end": '', "exam_list_start": '', "exam_list_open": '',"exam_list_kaowan":'', "course_info": course_info, 'notice': MoodleNotice.objects.all()})
            elif datetime.datetime.now() < user_obj.active_date:
                return render(request, 'moodle/player.html', {'user': user_obj, "video_list": video_list, "file_list": file_list, "exam_list_end": exam_list_end, "exam_list_start": exam_list_start, "exam_list_open": exam_list_open,"exam_list_kaowan":exam_list_kaowan, "course_info": course_info, 'notice': MoodleNotice.objects.all()})
            else:
                return render(request, 'moodle/player.html', {'user': user_obj, "video_list": '', "file_list": '', "exam_list_end": exam_list_end, "exam_list_start": exam_list_start, "exam_list_open": '',"exam_list_kaowan":exam_list_kaowan, "course_info": course_info, 'notice': MoodleNotice.objects.all()})
            # return render(request, 'moodle/player.html', {"video_list": video_list,  "exam_list_end": exam_list_end, "exam_list_start": exam_list_start, "exam_list_open": exam_list_open,"exam_list_kaowan":exam_list_kaowan, "course_info": course_info, 'notice': MoodleNotice.objects.all()})

        except Exception as e:
            print(3333333333333333333333333333,e)


class Excel(View):

    def get(self, request):

        return render(request, 'moodle/excel.html')

    def post(self, request):

        # print(MoodleCourse.objects.all())

        user_error = []
        course_error = []
        successful = ''

        # print(request.FILES['excel'])

        book = load_workbook(request.FILES['excel'])
        sheet = book.active

        for row in sheet.iter_rows(min_row=2):
            user_name = row[0].value
            m_ac = row[1].value
            # CUS = MoodleCUS()

            try:
                # 小程序有数据
                try:
                    user_obj = Users.objects.get(username=user_name, firstEmail=m_ac,isActive='true')
                except:
                    # print(row[1].value)
                    user_obj = Users.objects.get(username=user_name, secondEmail=m_ac,isActive='true')
                    m_ac = user_obj.firstEmail
                    # print(row[1].value,user_obj)

                # 判断moodle有没有数据
                if len(MoodleUser.objects.filter(m_ac=m_ac)) == 0:
                    # 不存在创建数据
                    new_m_user = MoodleUser()
                    new_m_user.m_ac = user_obj.firstEmail
                    h = hmac.new(salt, user_obj.domesticTelephone.encode(), digestmod='sha256')
                    new_m_ps_h = h.hexdigest()

                    new_m_user.m_ps = new_m_ps_h
                    new_m_user.users = user_obj
                    new_m_user.nick = user_obj.username
                    new_m_user.save()
                # else:
                #     old_moodle_user = MoodleUser.objects.filter(m_ac=m_ac)[0]
                #     old_moodle_user.active_date =

                for course in row[2::1]:

                    if course.value:
                        m_course = MoodleCourse.objects.filter(c_name__contains=course.value)

                        if len(m_course) == 1:

                            m_user = MoodleUser.objects.get(m_ac=m_ac)
                            if MoodleCUS.objects.filter(moodle_user=m_user.id, moodle_course=m_course[0].id):
                                pass
                            else:
                                CUS = MoodleCUS()
                                CUS.moodle_user = m_user
                                CUS.moodle_course = m_course[0]
                                CUS.save()

                                successful = 'OK'

                            # print(course.value, '存在导入成功')
                        else:
                            course_error.append((row[0].value, course.value))


            except Exception as e:
                # if len(Users.objects.filter(username=user_name)) > 0 :
                #
                #     print(e , user_name , m_ac , Users.objects.filter(username=user_name)[0].firstEmail)

                print(e , user_name , m_ac , Users.objects.filter(username=user_name))
                user_error.append(user_name)

        # print('用户存在列表', user_error)
        # print('课程名有误列表',course_error)

        if not (user_error or course_error):
            successful = 'OK'

        return render(request, 'moodle/excel.html', {'stu_error': user_error, 'course_error': course_error,'successful':successful})



def prof_course(request, prof_id=None):
    # print(webopenid)
    if request.method == 'GET':

        html = '<select name="moodle_course" required="" id="id_moodle_course"><option value="" selected="">---------</option>'

        if prof_id:

            prof_obj = MoodleUser.objects.get(id=prof_id)

            course_list = prof_obj.m_course.all()

            for cou in course_list:

                html += '<option value="{}">{}</option>'.format(cou.id,cou.c_name)

            html += '</select>'

        return HttpResponse(html)



