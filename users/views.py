# from django.http import JsonResponse
import json
import datetime
import os
import openpyxl

import requests
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect

from course.models import State, Course
from users.models import Users, School, Counselor, Assistant
from financial.models import Salary, Base_commission, Ambassador_annual, Head_salary
from moodle.models import MoodleCourse , MoodleVideo , MoodleSubject

from tool.openid_decorator import openid_check
from tool.orm_helper import data_generate, request_data_generate, field_data_generate, search_base_generate, \
    fuzzy_query_generate, search_ds_generate, fuzzy_obj

from tool.key import *

from pypinyin import lazy_pinyin

import requests
from lxml.html import etree
import time
import random


# Create your views here.
def index(request, webopenid=None):
    # print('OK')

    webopenid = request.GET.get("webOpenid", "")


    jscode = request.GET.get("code", "")
    url = base_url + "?appid=" + appid + "&secret=" + secret + "&js_code=" + jscode + "&grant_type=authorization_code"
    res = requests.get(url)

    # print('xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx',res.json())
    openid = res.json()['openid']

    if len(Users.objects.filter(openid=openid)) == 0:

        if webopenid:
            web_obj = Users.objects.get(webopenid=webopenid)
            web_obj.openid = openid
            web_obj.save()
        else:
            add = Users()
            add.role = 'xy'
            add.openid = openid
            add.save()

        return HttpResponse(json.dumps({"openid": openid}))
    else:
        return HttpResponse(json.dumps({"openid": openid}))


def web_index(request, webopenid=None):
    # print(webopenid)
    if request.method == 'GET':
        # print(request.GET)
        CODE = request.GET['code']
        # print(CODE)
        APPID = ''
        SECRET = ''
        webopenid = requests.get(
            '' + APPID + '&secret=' + SECRET + '&code=' + CODE + '&grant_type=authorization_code')
        webopenid = webopenid.json()['openid']
        # print(webopenid)

        # webopenid = '1235556789'
        use_obj = Users.objects.filter(webopenid=webopenid)
        if len(use_obj) == 0:
            # webopenid 不存在
            add = Users()
            add.role = 'xy'
            add.webopenid = webopenid
            add.save()
            sb = Users.objects.get(webopenid=webopenid)
            return render(request, 'web_index.html', {'webopenid': [webopenid], 'id': sb.id})
        else:
            # webopenid 存在
            webobj = Users.objects.get(webopenid=webopenid)
            if not webobj.isActive:
                return render(request, 'web_index.html', {'webopenid': [webopenid], 'id': webobj.id})

            if webobj.openid:
                # 且openid也存在 显示个人信息
                # 判断是否审核通过
                if webobj.isActive == 'false':
                    return render(request, 'web_result.html', {'reason': webobj.reason})
                    # 未通过
                elif webobj.isActive == 'new':
                    return render(request, 'web_wait.html')
                    # 审核中
                else:
                    # 显示个人信息
                    course_wk = set()
                    course_wk_list = webobj.course_set.filter(category='wk')
                    for i in course_wk_list:
                        course_wk.add(i.course)
                    course_wk = list(course_wk)
                    # 面授
                    course_ms = set()
                    course_ms_list = webobj.course_set.filter(category='ms')
                    for i in course_ms_list:
                        course_ms.add(i.course)
                    course_ms = list(course_ms)
                    if str(webobj.gender) == 'm':
                        gender_data = '男'
                    else:
                        gender_data = '女'

                    school = str(webobj.school.filter(func='zs')[0]) if len(
                        webobj.school.filter(func='zs')) != 0 else ''

                    return render(request, 'web_info.html', {
                        'username': str(webobj.username),
                        'gender': gender_data,
                        'birth': str(webobj.birth),
                        'IDCard': str(webobj.IDCard),
                        'major': str(webobj.major),
                        'domesticTelephone': str(webobj.domesticTelephone),
                        'foreignTelephone': str(webobj.foreignTelephone),
                        'domesticAddress': str(webobj.domesticAddress),
                        'foreignAddress': str(webobj.foreignAddress),
                        'ambassador': str(webobj.ambassador),
                        'paid': str(webobj.paid),
                        'count_wk': str(webobj.count_wk),
                        'count_ms': str(webobj.count_ms),
                        'balance': str(webobj.balance),
                        'course_wk': course_wk,
                        'course_ms': course_ms,
                        'school': school,
                        'grade': str(webobj.grade),
                        'firstEmail': str(webobj.firstEmail),
                        'secondEmail': str(webobj.secondEmail),
                        'wechat': str(webobj.wechat),
                    })

            else:
                # openid不存在 显示绑定小程序页面
                return render(request, 'web_bangding.html', {'webopenid': [webopenid]})
    if request.method == 'POST':
        webopenid_obj = Users.objects.get(webopenid=webopenid)
        school_name = request.POST.get("DXpicker", "")
        if request.POST.get("gender", "") == 'true':
            gender = 'm'
        else:
            gender = 'w'
        school_obj = School.objects.filter(s_name=school_name)
        webopenid_obj.username = request.POST.get("username", "")
        webopenid_obj.wechat = request.POST.get("wechat", "")
        webopenid_obj.GPA = request.POST.get("GPA", '')
        webopenid_obj.gender = gender
        webopenid_obj.birth = request.POST.get("birth", "")
        webopenid_obj.IDCard = '[' + request.POST.get("IDcard", "") + ']'
        webopenid_obj.major = request.POST.get("major", "")
        webopenid_obj.grade = request.POST.get("grade", "")
        webopenid_obj.school.set(school_obj)
        webopenid_obj.domesticTelephone = request.POST.get("tel_c", "")
        webopenid_obj.foreignTelephone = request.POST.get("tel_f", "")
        webopenid_obj.domesticAddress = request.POST.get("add_c", "")
        webopenid_obj.foreignAddress = request.POST.get("add_f", "")
        webopenid_obj.firstEmail = request.POST.get("email1", "")
        webopenid_obj.secondEmail = request.POST.get("email2", "")
        webopenid_obj.pinyin = request.POST.get("pinyin", "")
        webopenid_obj.study = request.POST.get("study", "")
        # 2020.1.13 修改 关闭财务审核功能
        # webopenid_obj.isActive = 'new'
        webopenid_obj.isActive = 'true'
        webopenid_obj.save()
        return render(request, 'web_bangding.html')


# def cw_index(request):
#     # if not mima:
#     #     return render(request,'cw_reg.html')
#     # if mima == 'Cso2019!':
#     #     return render(request,'cw_bangding.html')
#     # else:
#     #     return render(request,'cw_error.html',{'code':1})
#     if request.method == 'GET':
#         return render(request, 'cw_reg.html')
#     if request.method == 'POST':
#         try:
#             print(request.POST['password'])
#             if request.POST['password'] == 'Cso2019!':
#                 return render(request,'cw_bangding.html')
#             else:
#                 return render(request, 'cw_error.html', {'code': 1})
#         except:
#             return render(request,'cw_error.html',{'code':1})


def drop(request):
    if request.GET.keys():
        # 判断查询字符串
        for key in request.GET.keys():
            # print(key)
            if key == 'course':
                # 学校选择
                city_dict = {}

                CAN_Qset = School.objects.filter(func='zs',city='加拿大')
                USA_Qset = School.objects.filter(func='zs',city='美国')

                # for CAN in CAN_list:
                #     CAN_list.append(CAN.s_name)

                CAN_list = [ can.s_name for can in CAN_Qset ]
                USA_list = [ usa.s_name for usa in USA_Qset ]





                CAN_list = sorted(CAN_list, key=lambda ch: lazy_pinyin(ch))
                USA_list = sorted(USA_list, key=lambda ch: lazy_pinyin(ch))
                city_dict['加拿大'] = CAN_list
                city_dict['美国'] = USA_list

                result = {'code': 200, 'data': city_dict}
                return JsonResponse(result)

            if key == 'ambassador':
                ambassador_list = []
                ambassador_obj = Users.objects.filter(Q(role='ds') | Q(role='scfz') | Q(role='jqfz') | Q(role='qqfz'))
                for i in ambassador_obj:
                    ambassador_list.append(i.username)
                result = {'code': 200, 'data': ambassador_list}
                return JsonResponse(result)
            if key == 'qyfz':
                ambassador_list = []
                ambassador_obj = Users.objects.filter(Q(role='jqfz') | Q(role='qqfz'))
                for i in ambassador_obj:
                    ambassador_list.append(i.username)
                result = {'code': 200, 'data': ambassador_list}
                return JsonResponse(result)
            if key == 'scfz':
                ambassador_list = []
                ambassador_obj = Users.objects.filter(Q(role='scfz') | Q(role='jqfz') | Q(role='qqfz'))
                for i in ambassador_obj:
                    ambassador_list.append(i.username)
                result = {'code': 200, 'data': ambassador_list}
                return JsonResponse(result)
            if key == 'head':
                ds_name = request.GET['head']
                ds_obj = Users.objects.get(username=ds_name)
                ds_head = ds_obj.ambassador
                result = {'code': 200, 'data': str(ds_head)}
                return JsonResponse(result)

            else:
                result = {'code': 4001, 'error': 'Not enough permissions'}
                return JsonResponse(result)
    else:
        result = {'code': 4002, 'error': 'Not enough permissions'}
        return JsonResponse(result)


def users(request, openid=None, ):
    if request.method == 'GET':
        # print(openid)
        # 判断openid数据库是否存在
        try:
            openid_obj = Users.objects.get(openid=openid)
        except:
            result = {'code': 4003, 'error': 'Not enough permissions'}
            return JsonResponse(result)
        # print(openid_obj.role,22222222222222222)

        # ------------------------------ ▽ 自己基本数据▽ ----------------------------------------------
        # 网课
        course_wk = set()
        course_wk_list = openid_obj.course_set.filter(category='wk')
        for i in course_wk_list:
            course_wk.add(i.course)
        course_wk = list(course_wk)
        # 面授
        course_ms = set()
        course_ms_list = openid_obj.course_set.filter(category='ms')
        for i in course_ms_list:
            course_ms.add(i.course)
        course_ms = list(course_ms)

        if str(openid_obj.gender) == 'm':
            gender_data = '男'
        else:
            gender_data = '女'

        school = str(openid_obj.school.filter(func='zs')[0]) if len(openid_obj.school.filter(func='zs')) != 0 else ''

        data = {
            'username': str(openid_obj.username),
            'gender': gender_data,
            'GPA': str(openid_obj.GPA),
            'birth': str(openid_obj.birth),
            'IDCard': str(openid_obj.IDCard),
            'major': str(openid_obj.major),
            'domesticTelephone': str(openid_obj.domesticTelephone),
            'foreignTelephone': str(openid_obj.foreignTelephone),
            'domesticAddress': str(openid_obj.domesticAddress),
            'foreignAddress': str(openid_obj.foreignAddress),
            'ambassador': str(openid_obj.ambassador),
            'paid': str(openid_obj.paid),
            'count_wk': str(openid_obj.count_wk),
            'count_ms': str(openid_obj.count_ms),
            'balance': str(openid_obj.balance),
            'course_wk': course_wk,
            'course_ms': course_ms,
            'school': school,
            'grade': str(openid_obj.grade),
            'firstEmail': str(openid_obj.firstEmail),
            'secondEmail': str(openid_obj.secondEmail),
            'wechat': str(openid_obj.wechat),
            # 'role' : str(openid_obj.role),
            # 'isActive' : str(openid_obj.isActive),
        }
        # ------------------------------ △ 自己基本数据 △ ----------------------------------------------

        # 判断身份权限( xy / ds / fz / cw / zs )
        if openid_obj.role == 'xy':
            # print('111111111111111111111111111')
            # 判断学员当前账户状态( new / true / nds / false )
            if not openid_obj.isActive:
                # 不存在返回状态码201 前端显示学员信息填写界面
                result = {'code': 201, 'state': 'unregistered', 'data': data}
                return JsonResponse(result)
            if openid_obj.isActive == 'new':
                result = {'code': 202, 'reason': 2, 'state': 'In the review', 'data': data}
                return JsonResponse(result)

            if openid_obj.isActive in ('true', 'nds', 'update'):
                # print(openid_obj.isActive)
                # print(1)
                # try:
                #     if list(request.GET.keys())[0] == 'update':
                #         openid_obj.isActive = 'update'
                #         openid_obj.save()
                #         result = {'code': 200,'reason': 0}
                #         return JsonResponse(result)
                #     else:
                #         result = {'code': 200,'reason': 0, 'data': data}
                #         return JsonResponse(result)
                # except:
                #     result = {'code': 200,'reason': 0, 'data': data}
                #     return JsonResponse(result)
                result = {'code': 200, 'reason': 0, 'data': data}
                return JsonResponse(result)

            if openid_obj.isActive == 'false':
                try:
                    list_path = os.listdir('./media/{}/voucher'.format(openid_obj.id))
                except:
                    list_path = ''
                result = {'code': 203, 'reason': str(openid_obj.reason), 'path': list_path,
                          'id': str(openid_obj.id), 'stu_num': str(openid_obj.stu_num), 'data': data}

                return JsonResponse(result)

        # print(openid_obj.role)
        if openid_obj.role == 'ds':
            # 判断查询字符串
            # print(1)
            if request.GET.keys():

                key = list(request.GET.keys())[0]
                # 有查询字符串,根据key返回要查询的数据
                if key == 'my_stus':
                    # 查询邀请的正常学员
                    data = data_generate('Users', "ambassador={}, isActive='true'".format(openid_obj.id),
                                         ['username', 'id', 'stu_num'])
                    # print(data)
                    if type(data) is list:
                        result = {'code': 230, 'data': data}
                        return JsonResponse(result)
                    else:
                        data = [data]
                        result = {'code': 230, 'data': data}
                        return JsonResponse(result)

                if key == 'refund_stus':
                    # 查询邀请但已退费的学员
                    data = data_generate('Users', "ambassador={}, isActive='false'".format(openid_obj.id),
                                         ['username', 'id', 'stu_num'])
                    if type(data) is list:
                        result = {'code': 230, 'data': data}
                        return JsonResponse(result)
                    else:
                        data = [data]
                        result = {'code': 230, 'data': data}
                        return JsonResponse(result)

                if key == 'search':
                    # 格式为 ?search=查询内容
                    # data为满足条件对象的 id 姓名
                    search_msg = request.GET[key]
                    data = search_base_generate('Users', search_msg, ['username', 'id', 'stu_num'], openid_obj.id)
                    result = {'code': 210, 'data': data}
                    return JsonResponse(result)

                if key == 'bm':
                    # print(openid_obj.isActive,openid_obj.username)
                    if not openid_obj.isActive:
                        # 不存在返回状态码201 前端显示学员信息填写界面
                        result = {'code': 201, 'state': 'unregistered'}
                        return JsonResponse(result)

                    if openid_obj.isActive == 'false':
                        result = {'code': 203, 'reason': str(openid_obj.reason)}
                        return JsonResponse(result)

                    if openid_obj.isActive == 'new':
                        result = {'code': 202, 'state': 'In the review'}
                        return JsonResponse(result)

                    if openid_obj.isActive in ('true', 'nds', 'update'):
                        # print(openid_obj.isActive)
                        # print(1)
                        # try:
                        #     if list(request.GET.keys())[0] == 'update':
                        #         openid_obj.isActive = 'update'
                        #         openid_obj.save()
                        #         result = {'code': 200}
                        #         return JsonResponse(result)
                        #     else:
                        #         result = {'code': 200, 'data': data}
                        #         return JsonResponse(result)
                        # except:
                        #     result = {'code': 200, 'data': data}
                        #     return JsonResponse(result)
                        result = {'code': 200, 'data': data}
                        return JsonResponse(result)

                # key为ID 判断该ID的大使是否是当前用户
                try:
                    # print(key,openid_obj.id)
                    stu_obj = Users.objects.filter(id=key, ambassador=openid_obj.id)[0]
                    # print(stu_obj)
                except:
                    # 不是返回 405 权限不足
                    result = {'code': 4004, 'error': 'Not enough permissions'}
                    return JsonResponse(result)

                gen = str(stu_obj.gender)
                if gen == 'w':
                    gen = '女'
                else:
                    gen = '男'

                # 是则返回该ID 的基本数据
                data = {
                    'username': str(stu_obj.username),
                    'gender': gen,
                    'GPA': str(stu_obj.GPA),
                    'birth': str(stu_obj.birth)[0:10:1],
                    'IDCard': str(stu_obj.IDCard),
                    'major': str(stu_obj.major),
                    'domesticTelephone': str(stu_obj.domesticTelephone),
                    'foreignTelephone': str(stu_obj.foreignTelephone),
                    'domesticAddress': str(stu_obj.domesticAddress),
                    'foreignAddress': str(stu_obj.foreignAddress),
                    'ambassador': str(stu_obj.ambassador),
                    'paid': str(stu_obj.paid),
                    'count_wk': str(stu_obj.count_wk),
                    'count_ms': str(stu_obj.count_ms),
                    'balance': str(stu_obj.balance),
                    'sub_wk': data_generate('Users', 'id=' + key, ['sub_wk'])['sub_wk'],
                    'sub_ms': data_generate('Users', 'id=' + key, ['sub_ms'])['sub_ms'],
                    'reason': str(stu_obj.reason),
                    'refund_time': str(stu_obj.refund_time),
                    'wechat': str(stu_obj.wechat),
                }
                result = {'code': 210, 'data': data}
                return JsonResponse(result)

            else:
                # 无查询字符串则返回自己的数据
                result = {'code': 210, 'data': data}
                return JsonResponse(result)

        if openid_obj.role in ('scfz', 'jqfz', 'qqfz'):

            if request.GET.keys():
                key = list(request.GET.keys())[0]
                # 有查询字符串,根据key返回要查询的数据
                if key == 'my_stus':
                    # 查询邀请的正常学员
                    data = data_generate('Users', "ambassador={}, isActive='true'".format(openid_obj.id),
                                         ['username', 'id', 'stu_num'])
                    if type(data) is list:
                        result = {'code': 230, 'data': data}
                        return JsonResponse(result)
                    else:
                        data = [data]
                        result = {'code': 230, 'data': data}
                        return JsonResponse(result)

                if key == 'my_ambs':
                    # 负责人查询旗下大使
                    data = data_generate('Users', "Q(role='ds')|Q(role='scfz'),ambassador={}, isActive='true'".format(
                        openid_obj.id), ['username', 'id', 'stu_num'])
                    if type(data) is list:
                        result = {'code': 230, 'data': data}
                        return JsonResponse(result)
                    else:
                        data = [data]
                        result = {'code': 230, 'data': data}
                        return JsonResponse(result)

                if key == 'refund_stus':
                    # 查询邀请但已退费的学员
                    data = data_generate('Users', "ambassador={}, isActive='false'".format(openid_obj.id),
                                         ['username', 'id', 'stu_num'])
                    if type(data) is list:
                        result = {'code': 230, 'data': data}
                        return JsonResponse(result)
                    else:
                        data = [data]
                        result = {'code': 230, 'data': data}
                        return JsonResponse(result)

                if key == 'search':
                    # 格式为 ?search=查询内容
                    # data为满足条件对象的 id 姓名
                    search_msg = request.GET[key]
                    data = search_base_generate('Users', search_msg, ['username', 'id', 'stu_num'], openid_obj.id)
                    result = {'code': 220, 'data': data}
                    return JsonResponse(result)

                if key == 'search_ds':
                    # 格式为 ?search=查询内容
                    # data为满足条件对象的 id 姓名
                    search_msg = request.GET[key]
                    data = search_ds_generate('Users', search_msg, ['username', 'id', 'stu_num'], openid_obj.id)
                    result = {'code': 220, 'data': data}
                    return JsonResponse(result)

                if key == 'bm':
                    if not openid_obj.isActive:
                        # 不存在返回状态码201 前端显示学员信息填写界面
                        result = {'code': 201, 'state': 'unregistered'}
                        return JsonResponse(result)

                    # print(1)
                    if openid_obj.isActive == 'new':
                        result = {'code': 202, 'state': 'In the review'}
                        return JsonResponse(result)

                    if openid_obj.isActive in ('true', 'nds', 'update'):
                        # print(openid_obj.isActive)
                        # print(1)
                        # try:
                        #     if list(request.GET.keys())[0] == 'update':
                        #         openid_obj.isActive = 'update'
                        #         openid_obj.save()
                        #         result = {'code': 200}
                        #         return JsonResponse(result)
                        #     else:
                        #         result = {'code': 200, 'data': data}
                        #         return JsonResponse(result)
                        # except:
                        #     result = {'code': 200, 'data': data}
                        #     return JsonResponse(result)
                        result = {'code': 200, 'data': data}
                        return JsonResponse(result)

                if key == 'amb_smg':
                    value = request.GET[key]

                    try:
                        # print(value,openid_obj.id)
                        amb_obj = Users.objects.filter(id=value, ambassador=openid_obj.id)[0]
                        # print(amb_obj)
                    except:
                        # 不是返回 4005 权限不足
                        # print(1)
                        result = {'code': 4005, 'error': 'Not enough permissions'}
                        return JsonResponse(result)

                    if amb_obj.role in ('ds', 'scfz', 'jqfz', 'qqfz'):

                        now_month = datetime.datetime.now().month
                        year = datetime.datetime.now().year

                        if now_month == 1:
                            last_month = 12
                            year -= 1
                        else:
                            last_month = now_month - 1

                        try:
                            # print(amb_obj.username, '*********************')
                            salary_obj = Salary.objects.get(ds_name=amb_obj.username,
                                                            month=datetime.date(year, last_month, 1))

                            if str(amb_obj.gender) == 'm':
                                gender_data = '男'
                            else:
                                gender_data = '女'

                            sq_stu = [i.username for i in amb_obj.qxxy.filter(payStates='sq')]
                            bq_stu = [i.username for i in amb_obj.qxxy.filter(payStates='bq')]
                            qe_stu = [i.username for i in amb_obj.qxxy.filter(payStates='qe')]

                            # print(str(amb_obj.role))
                            base_salary = Head_salary.objects.get(level=str(amb_obj.role)).base_salary

                            school = str(amb_obj.school.filter(func='zs')[0]) if len(
                                amb_obj.school.filter(func='zs')) != 0 else ''

                            bonus = salary_obj.quarter_commission + salary_obj.ds_year_commission if amb_obj.role in (
                                'ds', 'scfz') else salary_obj.quarter_commission + salary_obj.fz_year_commission

                            role_dic = {
                                'ds': '大使',
                                'scfz': '市场负责人',
                                'jqfz': '区域负责人',
                                'qqfz': '区域负责人',
                            }

                            data = {
                                'username': str(amb_obj.username),
                                'gender': gender_data,
                                'role': role_dic[str(amb_obj.role)],
                                'IDCard': str(amb_obj.IDCard),
                                'domesticTelephone': str(amb_obj.domesticTelephone),
                                'fzr': str(amb_obj.ambassador),
                                'school': school,
                                'sq_stu': sq_stu,
                                'bq_stu': bq_stu,
                                'qe_stu': qe_stu,
                                'base_salary': base_salary,
                                'apply_commission': salary_obj.apply_commission,
                                'supplement_commission': salary_obj.supplement_commission,
                                'all_commission': salary_obj.all_commission,
                                'underling_commission': salary_obj.underling_commission,
                                'bonus': bonus,
                                'deduct': salary_obj.deduct,

                            }

                            result = {'code': 230, 'data': data}
                            return JsonResponse(result)

                        except:
                            if str(amb_obj.gender) == 'm':
                                gender_data = '男'
                            else:
                                gender_data = '女'

                            sq_stu = [i.username for i in amb_obj.qxxy.filter(payStates='sq')]
                            bq_stu = [i.username for i in amb_obj.qxxy.filter(payStates='bq')]
                            qe_stu = [i.username for i in amb_obj.qxxy.filter(payStates='qe')]

                            # print(str(amb_obj.role))
                            base_salary = Head_salary.objects.get(level=str(amb_obj.role)).base_salary

                            school = str(amb_obj.school.filter(func='zs')[0]) if len(
                                amb_obj.school.filter(func='zs')) != 0 else ''

                            role_dic = {
                                'ds': '大使',
                                'scfz': '市场负责人',
                                'jqfz': '区域负责人',
                                'qqfz': '区域负责人',
                            }

                            data = {
                                'username': str(amb_obj.username),
                                'gender': gender_data,
                                'role': role_dic[str(amb_obj.role)],
                                'IDCard': str(amb_obj.IDCard),
                                'domesticTelephone': str(amb_obj.domesticTelephone),
                                'fzr': str(amb_obj.ambassador),
                                'school': school,
                                'sq_stu': sq_stu,
                                'bq_stu': bq_stu,
                                'qe_stu': qe_stu,
                                'base_salary': base_salary,
                                'apply_commission': '数据暂未统计',
                                'supplement_commission': '数据暂未统计',
                                'all_commission': '数据暂未统计',
                                'underling_commission': '数据暂未统计',
                                'bonus': '数据暂未统计',
                                'deduct': '数据暂未统计',
                            }

                            result = {'code': 230, 'data': data}
                            return JsonResponse(result)

                    else:
                        data = {
                            'username': str(amb_obj.username),
                            'gender': str(amb_obj.gender),
                            'GPA': str(amb_obj.GPA),
                            'birth': str(amb_obj.birth)[0:10:1],
                            'IDCard': str(amb_obj.IDCard),
                            'major': str(amb_obj.major),
                            'domesticTelephone': str(amb_obj.domesticTelephone),
                            'foreignTelephone': str(amb_obj.foreignTelephone),
                            'domesticAddress': str(amb_obj.domesticAddress),
                            'foreignAddress': str(amb_obj.foreignAddress),
                            'ambassador': str(amb_obj.ambassador),
                            'paid': str(amb_obj.paid),
                            'count_wk': str(amb_obj.count_wk),
                            'count_ms': str(amb_obj.count_ms),
                            'balance': str(amb_obj.balance),
                            'sub_wk': data_generate('Users', 'id=' + value, ['sub_wk'])['sub_wk'],
                            'sub_ms': data_generate('Users', 'id=' + value, ['sub_ms'])['sub_ms'],
                        }
                        result = {'code': 210, 'data': data}
                        return JsonResponse(result)

                # key为ID 判断该ID的大使是否是当前用户
                try:
                    # print(key,openid_obj.id)
                    stu_obj = Users.objects.filter(id=key, ambassador=openid_obj.id)[0]
                    # print(stu_obj)
                except:
                    # 不是返回 405 权限不足
                    result = {'code': 4006, 'error': 'Not enough permissions'}
                    return JsonResponse(result)

                # 是则返回该ID 的基本数据
                data = {
                    'username': str(stu_obj.username),
                    'gender': str(stu_obj.gender),
                    'GPA': str(stu_obj.GPA),
                    'birth': str(stu_obj.birth)[0:10:1],
                    'IDCard': str(stu_obj.IDCard),
                    'major': str(stu_obj.major),
                    'domesticTelephone': str(stu_obj.domesticTelephone),
                    'foreignTelephone': str(stu_obj.foreignTelephone),
                    'domesticAddress': str(stu_obj.domesticAddress),
                    'foreignAddress': str(stu_obj.foreignAddress),
                    'ambassador': str(stu_obj.ambassador),
                    'paid': str(stu_obj.paid),
                    'count_wk': str(stu_obj.count_wk),
                    'count_ms': str(stu_obj.count_ms),
                    'balance': str(stu_obj.balance),
                    'sub_wk': data_generate('Users', 'id=' + key, ['sub_wk'])['sub_wk'],
                    'sub_ms': data_generate('Users', 'id=' + key, ['sub_ms'])['sub_ms'],
                }
                result = {'code': 210, 'data': data}
                return JsonResponse(result)

            else:
                # 无查询字符串则返回自己的数据
                result = {'code': 220, 'data': data}
                return JsonResponse(result)

        if openid_obj.role == 'cw':
            # print(1)

            # ------------------------------ ▽ 查时间判断数据更新 ▽ -------------------------------------------

            now_month = datetime.datetime.now().month
            year = datetime.datetime.now().year

            if now_month == 1:
                last_month = 12
                year -= 1
            else:
                last_month = now_month - 1

            if last_month == 1:
                last_last_month = 12
                last_month_year = year - 1
            else:
                last_last_month = last_month - 1
                last_month_year = year

            # 差工资表是否存在上月工资信息
            time = Salary.objects.filter(month=datetime.date(year, last_month, 1))
            # print(datetime.date(year, 9,30))
            if time:
                print('上月已计算')
            else:
                print('上月未计算,开始计算')
                ds_list = Users.objects.filter(role='ds')
                scfz_list = Users.objects.filter(role='scfz')
                jqfz_list = Users.objects.filter(role='jqfz')
                qqfz_list = Users.objects.filter(role='qqfz')
                qyfz_list = list(jqfz_list) + list(qqfz_list)

                # print(qyfz_list)

                def salary_generate(ds):
                    # 判断本月任务是否完成 未完成为0 完成为1
                    task = 0

                    # 上月的所有计算因数
                    last_month_factor = Base_commission.objects.get(month=last_month)
                    # 大使姓名
                    ds_name = ds.username

                    # ___________________________________________ 旗下申请费学员 ___________________________________________
                    # 上月交申请费人数
                    sq_num = len(ds.qxxy.filter(isActive='true', payStates='sq', pay_sq_time__year=year,
                                                pay_sq_time__month=last_month))
                    # print('申请费人数', sq_num)

                    # 上月份的申请费提成因数
                    apply = last_month_factor.application_fee

                    # 计算申请费提成
                    apply_commission = apply * sq_num
                    # print('申请费提成', apply_commission)

                    # ___________________________________________ 旗下补齐缴费学员 ___________________________________________
                    # 上月补齐缴费人数
                    bq_num = len(ds.qxxy.filter(isActive='true', payStates='bq', pay_sq_time__year=year,
                                                pay_sq_time__month=last_month))

                    # ___________________________________________ 自己补齐缴费 ___________________________________________

                    try:
                        if ds.payStates == 'bq' and ds.isActive == 'true' and ds.pay_sq_time__year == year and ds.pay_sq_time__month == last_month:
                            bq_num += 1
                    except Exception as e:
                        print(e)
                    # print('补齐缴费人数', bq_num)

                    # ___________________________________________ 旗下全额学员 ___________________________________________
                    # 上月全额缴费人数
                    qe_num = len(ds.qxxy.filter(isActive='true', payStates='qe', pay_qe_time__year=year,
                                                pay_qe_time__month=last_month))
                    try:
                        if ds.payStates == 'qe' and ds.isActive == 'true' and ds.pay_qe_time__year == year and ds.pay_qe_time__month == last_month:
                            bq_num += 1
                    except Exception as e:
                        print(e)

                    # print('全额缴费人数', ds.qxxy.filter(isActive='true', payStates='qe',pay_qe_time__year=year, pay_qe_time__month=last_month))

                    # 上月份人数目标
                    month_target = last_month_factor.month_target
                    # print('人数目标', month_target)

                    target_num = bq_num + qe_num
                    # print(bq_num,qe_num)
                    if target_num > month_target:
                        task = target_num - month_target
                        # 任务完成的提成因素
                        # 计算补缴提成
                        supplement_commission = last_month_factor.month_complete * bq_num

                        # 计算全额缴费提成
                        qe_commission = (apply + last_month_factor.month_complete) * qe_num

                        if bq_num > month_target:
                            supplement_commission -= month_target * 100

                        else:
                            supplement_commission -= bq_num * 100
                            qe_commission -= (month_target - bq_num) * 100

                        # print('补缴提成', supplement_commission)
                        # print('全额缴费提成', qe_commission)

                    else:
                        # 任务未完成的提成因素
                        # 计算补缴提成
                        supplement_commission = last_month_factor.month_unfinished * bq_num
                        # print('补缴提成', supplement_commission)

                        # 计算全额缴费提成
                        qe_commission = (apply + last_month_factor.month_unfinished) * qe_num
                        # print('全额缴费提成', qe_commission)

                    if ds.positive == '0' and target_num > 0:
                        if supplement_commission >= 100:
                            supplement_commission -= 100
                        else:
                            qe_commission -= 100

                        ds.positive = '1'
                        ds.save()

                    # ___________________________________________ 季度奖金 ___________________________________________

                    # 季度人数目标
                    quarter_target = last_month_factor.quarter_target
                    # if last_month == 8:  # 测试数据 正确数据在下面
                    if last_month == 9:
                        # print('第一季度')

                        start_date = datetime.date(year, 7, 1)
                        end_date = datetime.date(year, 9, 30)

                        # 季度补缴或全额人数
                        quarter_num = len(ds.qxxy.filter(Q(payStates='qe') | Q(payStates='bq'), isActive='true',
                                                         pay_sq_time__range=(start_date, end_date)))
                        # print('季度补缴或全额人数', quarter_num)

                        if quarter_num >= quarter_target:
                            quarter_commission = last_month_factor.quarter_complete
                        else:
                            quarter_commission = 0

                    elif last_month == 12:
                        # print('第二季度')

                        start_date = datetime.date(year, 10, 1)
                        end_date = datetime.date(year, 12, 31)

                        quarter_num = len(ds.qxxy.filter(Q(payStates='qe') | Q(payStates='bq'), isActive='true',
                                                         pay_sq_time__range=(start_date, end_date)))

                        if quarter_num >= quarter_target:
                            quarter_commission = last_month_factor.quarter_complete
                        else:
                            quarter_commission = 0

                    elif last_month == 4:
                        # print('第三季度')

                        start_date = datetime.date(year, 1, 1)
                        end_date = datetime.date(year, 4, 30)

                        quarter_num = len(ds.qxxy.filter(Q(payStates='qe') | Q(payStates='bq'), isActive='true',
                                                         pay_sq_time__range=(start_date, end_date)))

                        if quarter_num >= quarter_target:
                            quarter_commission = last_month_factor.quarter_complete
                        else:
                            quarter_commission = 0

                    else:
                        quarter_commission = 0

                    # print('季度奖金', quarter_commission)

                    now_time = str(year) + '-' + str(last_month).zfill(2) + '-01'
                    return ds_name, now_time, apply_commission, supplement_commission, qe_commission, quarter_commission, task

                # 区域负责人工资 ********************************************************************************************************************************************************

                for qyfz in qyfz_list:
                    # print(qyfz)
                    annual_bonus = 0
                    ds_name, now_time, apply_commission, supplement_commission, qe_commission, quarter_commission, task = salary_generate(
                        qyfz)
                    # print(ds_name, now_time, apply_commission, supplement_commission, qe_commission, quarter_commission)
                    # print(now_time)

                    # ___________________________________________ 年度奖金 ___________________________________________
                    # if last_month == 8:  # 测试数据 正确数据在下面
                    if last_month == 6:
                        school_list = School.objects.filter(Q(users__payStates='bq') | Q(users__payStates='qe'),
                                                            func='zs', qyfz=qyfz.username).distinct()
                        for sch in school_list:
                            stu_num = len(Users.objects.filter(school=sch))
                            # print(sch,'市场有',stu_num,'个学生')
                            # print(sch.market)
                            if sch.market == 'big':
                                if 70 <= stu_num <= 149:
                                    annual_bonus += sch.market.low
                                elif 150 <= stu_num <= 299:
                                    annual_bonus += sch.market.middle
                                elif 300 <= stu_num:
                                    annual_bonus += sch.market.high
                                else:
                                    annual_bonus += 0

                            elif sch.market == 'middle':
                                if 30 <= stu_num <= 69:
                                    annual_bonus += sch.market.low
                                elif 70 <= stu_num <= 99:
                                    annual_bonus += sch.market.middle
                                elif 100 <= stu_num:
                                    annual_bonus += sch.market.high
                                else:
                                    annual_bonus += 0

                            elif sch.market == 'small':
                                if 10 <= stu_num <= 29:
                                    annual_bonus += sch.market.low
                                elif 30 <= stu_num <= 69:
                                    annual_bonus += sch.market.middle
                                elif 70 <= stu_num:
                                    annual_bonus += sch.market.high
                                else:
                                    annual_bonus += 0

                            else:
                                result = {'code': 4007, 'error': 'Not enough permissions'}
                                return JsonResponse(result)

                    else:
                        annual_bonus = 0

                    # ___________________________________________ 下属提成 ___________________________________________
                    underling_commission = 0

                    # 旗下大使的提成.......................................................................
                    qxds_list = qyfz.qxxy.filter(role='ds')
                    # print('市场负责人名字',scfz,'旗下大使列表',qxds_list)
                    ds_underling_num = 0

                    for ds in qxds_list:
                        sq_num = len(ds.qxxy.filter(isActive='true', payStates='sq', pay_sq_time__year=year,
                                                    pay_sq_time__month=last_month))
                        qe_num = len(ds.qxxy.filter(isActive='true', payStates='qe', pay_qe_time__year=year,
                                                    pay_qe_time__month=last_month))
                        pay_num = sq_num + qe_num

                        # print(qyfz.single)
                        if qyfz.single == '0' and pay_num > 0:
                            ds_underling_num += 1
                            qyfz.single = 1
                            qyfz.save()

                        ds_underling_num += pay_num

                    underling_commission += ds_underling_num * 50

                    # 旗下市场负责人的提成.......................................................................
                    qxfz_list = qyfz.qxxy.filter(role='scfs')
                    # print('市场负责人名字',scfz,'旗下大使列表',qxds_list)
                    fz_underling_num = 0

                    for fz in qxfz_list:
                        sq_num = len(fz.qxxy.filter(isActive='true', payStates='sq', pay_sq_time__year=year,
                                                    pay_sq_time__month=last_month))
                        qe_num = len(fz.qxxy.filter(isActive='true', payStates='qe', pay_qe_time__year=year,
                                                    pay_qe_time__month=last_month))
                        pay_num = sq_num + qe_num

                        fz_underling_num += pay_num

                    underling_commission += fz_underling_num * 50

                    # 旗下市场负责人旗下的大使的提成.......................................................................
                    qxfz_list = qyfz.qxxy.filter(role='scfs')
                    # print('市场负责人名字',scfz,'旗下大使列表',qxds_list)
                    underling_num = 0

                    for fz in qxfz_list:
                        qxfz_qxds_list = fz.qxxy.filter(role='ds')

                        for fz_ds in qxfz_qxds_list:
                            sq_num = len(fz_ds.qxxy.filter(isActive='true', payStates='sq', pay_sq_time__year=year,
                                                           pay_sq_time__month=last_month))
                            qe_num = len(fz_ds.qxxy.filter(isActive='true', payStates='qe', pay_qe_time__year=year,
                                                           pay_qe_time__month=last_month))
                            pay_num = sq_num + qe_num

                            underling_num += pay_num

                        underling_commission += underling_num * 50

                    # ___________________________________________ 退费计算 ___________________________________________

                    deduct_num = len(qyfz.qxxy.filter(
                        isActive='false', payStates='yt', refund_time__year=last_month_year,
                        refund_time__month=last_last_month))

                    try:
                        task = Salary.objects.get(month__year=last_month_year, month__month=last_last_month)

                        if task != 0:
                            # 上上月的任务完成了 判断退费人数和上上月完成任务超出人数差
                            if deduct_num < task:
                                deduct = deduct_num * 250
                            else:
                                deduct = (deduct_num - task) * 150 + task * 250
                        else:
                            # 上上月任务未完成 扣150
                            deduct = deduct_num * 150
                    except:
                        deduct = 0

                    # ___________________________________________ 数据写入 ___________________________________________

                    if qyfz.role == 'qqfz':
                        base_salary = 3000
                    elif qyfz.role == 'jqfz':
                        base_salary = 600
                    else:
                        base_salary = 0

                    salary = Salary()
                    salary.head = qyfz
                    salary.ds_name = ds_name
                    salary.month = now_time
                    salary.base_salary = base_salary
                    salary.apply_commission = apply_commission
                    salary.supplement_commission = supplement_commission
                    salary.all_commission = qe_commission
                    salary.quarter_commission = quarter_commission
                    salary.fz_year_commission = annual_bonus
                    salary.underling_commission = underling_commission
                    salary.deduct = deduct
                    salary.task = task

                    salary.save()

                # 市场负责人工资 ********************************************************************************************************************************************************
                for scfz in scfz_list:
                    ds_name, now_time, apply_commission, supplement_commission, qe_commission, quarter_commission, task = salary_generate(
                        scfz)

                    # ___________________________________________ 年度奖金 ___________________________________________
                    if last_month == 8:  # 测试数据 正确数据在下面
                        # if last_month == 6:
                        # print(year - 1, '年度')

                        start_date = datetime.date(year - 1, 7, 1)
                        end_date = datetime.date(year, 4, 30)
                        annual_qe_num = len(scfz.qxxy.filter(
                            Q(payStates='qe') & Q(pay_qe_time__range=(start_date, end_date), isActive='true')))
                        annual_bq_num = len(scfz.qxxy.filter(
                            Q(payStates='bq') & Q(pay_bq_time__range=(start_date, end_date), isActive='true')))
                        annual_num = annual_bq_num + annual_qe_num
                        # print('全额人数', annual_qe_num)
                        # print('补齐人数', annual_bq_num)

                        level = Ambassador_annual.objects.get(floor__lte=annual_num, ceiling__gte=annual_num)
                        annual_bonus = level.annual_bonus
                        discount = level.discount

                        # print('年度奖金', annual_bonus)
                        # print('学费减免资格', discount)
                    else:
                        annual_bonus = 0
                        discount = 0

                    # ___________________________________________ 下属提成 ___________________________________________
                    qxds_list = scfz.qxxy.filter(role='ds')
                    # print('市场负责人名字',scfz,'旗下大使列表',qxds_list)
                    underling_num = 0

                    for ds in qxds_list:
                        sq_num = len(ds.qxxy.filter(isActive='true', payStates='sq', pay_sq_time__year=year,
                                                    pay_sq_time__month=last_month))
                        qe_num = len(ds.qxxy.filter(isActive='true', payStates='qe', pay_qe_time__year=year,
                                                    pay_qe_time__month=last_month))
                        pay_num = sq_num + qe_num

                        # print(scfz.single)
                        if scfz.single == '0' and pay_num > 0:
                            underling_num += 1
                            scfz.single = 1
                            scfz.save()

                        underling_num += pay_num

                    underling_commission = underling_num * 50

                    # ___________________________________________ 退费计算 ___________________________________________

                    deduct_num = len(scfz.qxxy.filter(
                        isActive='false', payStates='yt', refund_time__year=last_month_year,
                        refund_time__month=last_last_month))

                    try:
                        task = Salary.objects.get(month__year=last_month_year, month__month=last_last_month)

                        if task != 0:
                            # 上上月的任务完成了 判断退费人数和上上月完成任务超出人数差
                            if deduct_num < task:
                                deduct = deduct_num * 250
                            else:
                                deduct = (deduct_num - task) * 150 + task * 250
                        else:
                            # 上上月任务未完成 扣150
                            deduct = deduct_num * 150
                    except:
                        deduct = 0

                    # ___________________________________________ 数据写入 ___________________________________________

                    salary = Salary()
                    salary.head = scfz
                    salary.ds_name = ds_name
                    salary.month = now_time
                    salary.base_salary = 0
                    salary.apply_commission = apply_commission
                    salary.supplement_commission = supplement_commission
                    salary.all_commission = qe_commission
                    salary.quarter_commission = quarter_commission
                    salary.fz_year_commission = annual_bonus
                    salary.discount = discount
                    salary.underling_commission = underling_commission
                    salary.deduct = deduct
                    salary.task = task

                    salary.save()

                    if scfz.ambassador:
                        try:
                            # print(scfz.username,'的上级是',scfz.ambassador.username)
                            scfz_sal = Salary.objects.get(ds_name=scfz.ambassador.username,
                                                          month=datetime.date(year, last_month, 1))
                            scfz_sal.deduct = scfz_sal.deduct + deduct_num * 50
                            scfz_sal.save()
                        except:
                            pass

                # 大使工资 ********************************************************************************************************************************************************
                for ds in ds_list:
                    ds_name, now_time, apply_commission, supplement_commission, qe_commission, quarter_commission, task = salary_generate(
                        ds)
                    # print(ds_name, now_time, apply_commission, supplement_commission, qe_commission, quarter_commission)
                    # print(now_time)

                    # ___________________________________________ 年度奖金 ___________________________________________
                    # if last_month == 8:  # 测试数据 正确数据在下面
                    if last_month == 6:
                        # print(year - 1, '年度')

                        start_date = datetime.date(year - 1, 7, 1)
                        end_date = datetime.date(year, 4, 30)
                        annual_qe_num = len(ds.qxxy.filter(
                            Q(payStates='qe') & Q(pay_qe_time__range=(start_date, end_date), isActive='true')))
                        annual_bq_num = len(ds.qxxy.filter(
                            Q(payStates='bq') & Q(pay_bq_time__range=(start_date, end_date), isActive='true')))
                        annual_num = annual_bq_num + annual_qe_num
                        # print('全额人数', annual_qe_num)
                        # print('补齐人数', annual_bq_num)

                        level = Ambassador_annual.objects.get(floor__lte=annual_num, ceiling__gte=annual_num)
                        annual_bonus = level.annual_bonus
                        discount = level.annual_bonus

                        # print('年度奖金', annual_bonus)
                        # print('学费减免资格', discount)

                    else:
                        annual_bonus = 0
                        discount = 0

                    # ___________________________________________ 退费计算 ___________________________________________

                    deduct_num = len(ds.qxxy.filter(
                        isActive='false', payStates='yt', refund_time__year=last_month_year,
                        refund_time__month=last_last_month))

                    try:
                        task = Salary.objects.get(month__year=last_month_year, month__month=last_last_month)

                        if task != 0:
                            # 上上月的任务完成了 判断退费人数和上上月完成任务超出人数差
                            if deduct_num < task:
                                deduct = deduct_num * 250
                            else:
                                deduct = (deduct_num - task) * 150 + task * 250
                        else:
                            # 上上月任务未完成 扣150
                            deduct = deduct_num * 150
                    except:
                        deduct = 0

                    # ___________________________________________ 数据写入 ___________________________________________

                    salary = Salary()
                    salary.head = ds
                    salary.ds_name = ds_name
                    salary.month = now_time
                    salary.apply_commission = apply_commission
                    salary.supplement_commission = supplement_commission
                    salary.all_commission = qe_commission
                    salary.quarter_commission = quarter_commission
                    salary.ds_year_commission = annual_bonus
                    salary.discount = discount
                    salary.deduct = deduct
                    salary.task = task

                    salary.save()

                    # 根据本月退费人数 所有上级退费 每有一人50刀
                    if ds.ambassador:
                        try:

                            # print(ds.username,'的上级是',ds.ambassador.username)
                            scfz_sal = Salary.objects.get(ds_name=ds.ambassador.username,
                                                          month=datetime.date(year, last_month, 1))
                            scfz_sal.deduct = scfz_sal.deduct + deduct_num * 50
                            scfz_sal.save()

                            if ds.ambassador.ambassador:
                                # print(ds.username,'的上上级是',ds.ambassador.ambassador.username)
                                qyfz_sal = Salary.objects.get(ds_name=ds.ambassador.ambassador.username,
                                                              month=datetime.date(year, last_month, 1))
                                qyfz_sal.deduct = qyfz_sal.deduct + deduct_num * 50
                                qyfz_sal.save()

                        except:
                            pass
                        # print(ds.username, '的退费人数是', deduct_num, '人')
                        # print('-------------------')

            # ------------------------------ △ 查时间判断数据更新 △ ------------------------------------------------

            stu_search_conditions = {}

            if request.GET.keys():

                # print(1)

                # 判断查询字符串
                for key in request.GET.keys():
                    # print(key)
                    if key == 'all_new':
                        # 根据前端查询字符串查询 新学员信息
                        data = request_data_generate('Users', request, "Q(isActive='new')|Q(isActive='update')", key)
                        if type(data) is list:
                            result = {'code': 230, 'data': data}
                            return JsonResponse(result)
                        else:
                            data = [data]
                            result = {'code': 230, 'data': data}
                            return JsonResponse(result)

                    elif key == 'all_stu':
                        # 根据前端查询字符串查询 所有学员信息
                        data = request_data_generate('Users', request, "isActive='true'", key)
                        if type(data) is list:
                            result = {'code': 230, 'data': data}
                            return JsonResponse(result)
                        else:
                            data = [data]
                            result = {'code': 230, 'data': data}
                            return JsonResponse(result)

                    elif key == 'new_ds':

                        # 根据前端查询字符串查询 所有学员信息
                        data = request_data_generate('Users', request, "isActive='nds'", key)
                        # print(type(data))
                        if type(data) is list:
                            result = {'code': 230, 'data': data}
                            return JsonResponse(result)
                        else:
                            data = [data]
                            result = {'code': 230, 'data': data}
                            return JsonResponse(result)

                    elif key == 'course_msg':
                        pass


                    # ------------------------------ ▽ 搜索学生信息 ▽ ----------------------------------------------
                    elif key == 'search_stu_pay':
                        value = request.GET[key]
                        if value == '申请费':
                            value = 'sq'
                        elif value == '补齐':
                            value = 'bq'
                        elif value == '全额':
                            value = 'qe'

                        stu_search_conditions[key] = value


                    elif key == 'search_stu_time':
                        value = request.GET[key]
                        if value == '':
                            pass
                        else:
                            value = str(value).zfill(2)
                        # print(value)
                        stu_search_conditions[key] = value


                    # url里这个必须有 且必须放在最后一位
                    elif key == 'search_stu_msg':

                        msg = request.GET[key]
                        try:
                            int(msg)
                            msg = "Q(id__contains='" + msg + "'),"
                        except:
                            msg = "Q(username__contains='" + msg + "'),"

                        pay = stu_search_conditions.get('search_stu_pay', '')
                        if pay:
                            pay = "payStates='" + pay + "',"

                        time = stu_search_conditions.get('search_stu_time', '')
                        if time:
                            time = "Q(pay_sq_time__contains='" + str(
                                datetime.datetime.now().year) + '-' + time + "')|Q(pay_bq_time__contains='" + str(
                                datetime.datetime.now().year) + '-' + time + "')|Q(pay_qe_time__contains='" + str(
                                datetime.datetime.now().year) + '-' + time + "'),"

                        # print("Users.objects.filter(" + msg + time + pay + "isActive='true')")
                        data = eval("Users.objects.filter(" + msg + time + pay + "~Q(wechat=''),isActive='true')")
                        data_list = []
                        for i in data:
                            data_set = {}
                            data_set['id'] = i.id
                            try:
                                data_set['school'] = School.objects.filter(func='zs', users__username=i.username)[
                                    0].s_name
                            except Exception as e:
                                print(e)
                            data_set['username'] = i.username
                            if i.currency == '加币':
                                data_set['pay'] = str(round(i.paid, 2)) + '＄'
                            else:
                                data_set['pay'] = str(round(i.paid * 5.2, 2)) + '￥'
                            data_list.append(data_set)

                        result = {'code': 230, 'data': data_list,
                                  'all_num': len(Users.objects.filter(~Q(username=''), ~Q(wechat=''), isActive='true')),
                                  'stu_num': len(data_list)}
                        return JsonResponse(result)


                    # ------------------------------ △ 搜索学生信息 △ ----------------------------------------------

                    elif key == 'search_amb':
                        value = request.GET[key]
                        try:
                            value = int(value)
                            data = data_generate('Users',
                                                 "Q(role='ds')|Q(role='scfz')|Q(role='jqfz')|Q(role='qqfz'),id__contains='{}'".format(
                                                     value), ['id', 'username', 'role', 'school_zs'])
                            # print(data)
                            if type(data) is list:
                                result = {'code': 230, 'data': data, 'stu_num': len(data), 'all_num': len(
                                    Users.objects.filter(
                                        Q(role='ds') | Q(role='scfz') | Q(role='jqfz') | Q(role='qqfz')))}
                                return JsonResponse(result)
                            else:
                                data = [data]
                                result = {'code': 230, 'data': data, 'stu_num': len(data), 'all_num': len(
                                    Users.objects.filter(
                                        Q(role='ds') | Q(role='scfz') | Q(role='jqfz') | Q(role='qqfz')))}
                                return JsonResponse(result)
                        except:
                            data = data_generate('Users',
                                                 "Q(role='ds')|Q(role='scfz')|Q(role='jqfz')|Q(role='qqfz'),username__contains='{}'".format(
                                                     value),
                                                 ['id', 'username', 'role', 'school_zs'])
                            # print(data)
                            if type(data) is list:
                                result = {'code': 230, 'data': data, 'stu_num': len(data), 'all_num': len(
                                    Users.objects.filter(
                                        Q(role='ds') | Q(role='scfz') | Q(role='jqfz') | Q(role='qqfz')))}
                                return JsonResponse(result)
                            else:
                                data = [data]
                                result = {'code': 230, 'data': data, 'stu_num': len(data), 'all_num': len(
                                    Users.objects.filter(
                                        Q(role='ds') | Q(role='scfz') | Q(role='jqfz') | Q(role='qqfz')))}
                                return JsonResponse(result)


                    elif key == 'amb_smg':
                        if request.GET[key]:
                            value = request.GET[key]
                            amb_obj = Users.objects.get(id=value)
                            try:
                                salary_obj = Salary.objects.get(ds_name=amb_obj.username,
                                                                month=datetime.date(year, last_month, 1))
                                bonus = salary_obj.quarter_commission + salary_obj.ds_year_commission if amb_obj.role in (
                                    'ds', 'scfz') else salary_obj.quarter_commission + salary_obj.fz_year_commission

                                apply_commission = salary_obj.apply_commission
                                supplement_commission = salary_obj.supplement_commission
                                all_commission = salary_obj.all_commission
                                underling_commission = salary_obj.underling_commission
                                deduct = salary_obj.deduct
                            except:
                                bonus = '暂未计算'
                                apply_commission = '暂未计算'
                                supplement_commission = '暂未计算'
                                all_commission = '暂未计算'
                                underling_commission = '暂未计算'
                                deduct = '暂未计算'
                            if amb_obj.role in ('ds', 'scfz', 'jqfz', 'qqfz'):
                                if str(amb_obj.gender) == 'm':
                                    gender_data = '男'
                                else:
                                    gender_data = '女'

                                sq_stu = [i.username for i in amb_obj.qxxy.filter(payStates='sq')]
                                bq_stu = [i.username for i in amb_obj.qxxy.filter(payStates='bq')]
                                qe_stu = [i.username for i in amb_obj.qxxy.filter(payStates='qe')]
                                yt_stu = [i.username for i in amb_obj.qxxy.filter(payStates='yt')]

                                # print(str(amb_obj.role))
                                base_salary = Head_salary.objects.get(level=str(amb_obj.role)).base_salary

                                school = str(amb_obj.school.filter(func='zs')[0]) if len(
                                    amb_obj.school.filter(func='zs')) != 0 else ''

                                role_dic = {
                                    'ds': '大使',
                                    'scfz': '市场负责人',
                                    'jqfz': '区域负责人',
                                    'qqfz': '区域负责人',
                                }

                                data = {
                                    'username': str(amb_obj.username),
                                    'gender': gender_data,
                                    'role': role_dic[str(amb_obj.role)],
                                    'IDCard': str(amb_obj.IDCard),
                                    'domesticTelephone': str(amb_obj.domesticTelephone),
                                    'fzr': str(amb_obj.ambassador),
                                    'school': school,
                                    'sq_stu': sq_stu,
                                    'bq_stu': bq_stu,
                                    'qe_stu': qe_stu,
                                    'yt_stu': yt_stu,
                                    'base_salary': base_salary,
                                    'apply_commission': apply_commission,
                                    'supplement_commission': supplement_commission,
                                    'all_commission': all_commission,
                                    'underling_commission': underling_commission,
                                    'bonus': bonus,
                                    'deduct': deduct,

                                }

                                result = {'code': 230, 'data': data}
                                return JsonResponse(result)
                            else:
                                result = {'code': 4008, 'error': 'Not enough permissions'}
                                return JsonResponse(result)


                        else:
                            result = {'code': 4009, 'error': 'Not enough permissions'}
                            return JsonResponse(result)



                    # elif key == 'all_amb':
                    #     # 根据前端查询字符串查询 所有大使信息
                    #     # data = request_data_generate('Users', request, "role='ds'", key)
                    #     data = data_generate('Users', "role='ds'", ['username', 'school_sk','id'])
                    #     # print(data)
                    #     result = {'code': 230, 'data': data}
                    #     return JsonResponse(result)

                    elif key == 'all_ref':
                        # 根据前端查询字符串查询 所有要退费学员信息
                        data = request_data_generate('Users', request, "isActive='refund'", key)
                        if type(data) is list:
                            result = {'code': 230, 'data': data}
                            return JsonResponse(result)
                        else:
                            data = [data]
                            result = {'code': 230, 'data': data}
                            return JsonResponse(result)


                    elif key == 'img':
                        if request.GET[key]:
                            value = request.GET[key]
                            data = {}
                            try:
                                data['card_path'] = os.listdir('./media/{}/card_img'.format(value))
                            except:
                                data['card_path'] = ''
                            try:
                                data['voucher_path'] = os.listdir('./media/{}/voucher'.format(value))
                            except:
                                data['voucher_path'] = ''
                            try:
                                data['LOP_path'] = os.listdir('./media/{}/LOP_img'.format(value))
                            except:
                                data['LOP_path'] = ''
                            try:
                                data['GPA_path'] = os.listdir('./media/{}/GPA_img'.format(value))
                            except:
                                data['GPA_path'] = ''
                            result = {'code': 230, 'data': data}
                            return JsonResponse(result)
                        else:
                            result = {'code': 4010, 'error': 'Not enough permissions'}
                            return JsonResponse(result)




                    # 不是以上查询字符串 则只能是 id
                    else:

                        # ------------------------------ ▽ 学生基本数据▽ ----------------------------------------------
                        # 网课
                        try:
                            result_obj = Users.objects.get(id=key)
                        except:
                            result = {'code': 4011, 'error': 'Not enough permissions'}
                            return JsonResponse(result)

                        course_wk = set()
                        course_wk_list = result_obj.course_set.filter(category='wk')
                        for i in course_wk_list:
                            course_wk.add(i.course)
                        course_wk = list(course_wk)
                        # 面授
                        course_ms = set()
                        course_ms_list = result_obj.course_set.filter(category='ms')
                        for i in course_ms_list:
                            course_ms.add(i.course)
                        course_ms = list(course_ms)

                        if str(result_obj.gender) == 'm':
                            gender_data = '男'
                        else:
                            gender_data = '女'

                        school = str(result_obj.school.filter(func='zs')[0]) if len(
                            result_obj.school.filter(func='zs')) != 0 else ''

                        result_data = {
                            'username': str(result_obj.username),
                            'gender': gender_data,
                            'GPA': str(result_obj.GPA),
                            'birth': str(result_obj.birth),
                            'IDCard': str(result_obj.IDCard),
                            'major': str(result_obj.major),
                            'domesticTelephone': str(result_obj.domesticTelephone),
                            'foreignTelephone': str(result_obj.foreignTelephone),
                            'domesticAddress': str(result_obj.domesticAddress),
                            'foreignAddress': str(result_obj.foreignAddress),
                            'ambassador': str(result_obj.ambassador),
                            'paid': str(result_obj.paid),
                            'count_wk': str(result_obj.count_wk),
                            'count_ms': str(result_obj.count_ms),
                            'balance': str(result_obj.balance),
                            'course_wk': course_wk,
                            'course_ms': course_ms,
                            'school': school,
                            'grade': str(result_obj.grade),
                            'firstEmail': str(result_obj.firstEmail),
                            'secondEmail': str(result_obj.secondEmail),
                            'role': str(result_obj.role),
                            'wechat': str(result_obj.wechat),
                            'currency': str(result_obj.currency),
                            # 'isActive' : str(openid_obj.isActive),
                        }
                        # ------------------------------ △ 学生基本数据 △ ----------------------------------------------

                        try:
                            # data = request_data_generate('Users', request, "id=" + key, key)
                            result = {'code': 230, 'data': result_data}
                            return JsonResponse(result)
                        except:
                            # 查询字符串也不是id 返回405 权限不足
                            result = {'code': 4012, 'error': 'Not enough permissions'}
                            return JsonResponse(result)

            # 无查询字符串返回 新学员/退款申请学员/新大使 数量
            else:
                news = Users.objects.filter(Q(isActive='new') | Q(isActive='update'))
                refunds = Users.objects.filter(isActive='refund')
                nds = Users.objects.filter(isActive='nds')
                result = {'code': 230, 'data': {'news': len(news), 'refunds': len(refunds), 'nds': len(nds)}}
                return JsonResponse(result)

        # if openid_obj.role == 'zs':

    if request.method == 'POST':

        if openid:
            openid_obj = Users.objects.get(openid=openid)

            if openid_obj.role in ('', 'xy', 'ds', 'scfz', 'qqfz', 'jqfz'):
                # print(99999)
                user_type = request.POST.get("type", "")
                if user_type == 'ds':
                    # print(9999)
                    if request.POST.get("gender", "") == 'true':
                        gender = 'm'
                    else:
                        gender = 'w'
                    # school_name = request.POST.get("DXpicker", "")
                    # school_obj = School.objects.filter(s_name=school_name)
                    # openid_obj.school.set(school_obj)
                    openid_obj.username = request.POST.get('username')
                    openid_obj.gender = gender
                    openid_obj.domesticTelephone = request.POST.get('domesticTelephone')
                    # openid_obj.pinyin = request.POST.get('pinyin')
                    # openid_obj.study = request.POST.get('study')
                    openid_obj.IDCard = '[' + request.POST.get("IDCard", "") + ']'
                    # print(request.POST.get("IDCard", ""))
                    openid_obj.role = 'xy'
                    openid_obj.isActive = 'nds'
                    openid_obj.save()
                else:
                    # print(999)
                    school_name = request.POST.get("DXpicker", "")
                    if request.POST.get("gender", "") == 'true':
                        gender = 'm'
                    else:
                        gender = 'w'
                    school_obj = School.objects.filter(s_name=school_name)
                    openid_obj.username = request.POST.get("username", "")
                    openid_obj.wechat = request.POST.get("wechat", "")
                    openid_obj.GPA = request.POST.get("GPA", '')
                    openid_obj.gender = gender
                    openid_obj.birth = request.POST.get("birth", "")
                    openid_obj.IDCard = '[' + request.POST.get("IDcard", "") + ']'
                    openid_obj.major = request.POST.get("major", "")
                    openid_obj.grade = request.POST.get("grade", "")
                    openid_obj.school.set(school_obj)
                    openid_obj.domesticTelephone = request.POST.get("tel_c", "")
                    openid_obj.foreignTelephone = request.POST.get("tel_f", "")
                    openid_obj.domesticAddress = request.POST.get("add_c", "")
                    openid_obj.foreignAddress = request.POST.get("add_f", "")
                    openid_obj.firstEmail = request.POST.get("email1", "")
                    openid_obj.secondEmail = request.POST.get("email2", "")
                    openid_obj.pinyin = request.POST.get("pinyin", "")
                    openid_obj.study = request.POST.get("study", "")
                    openid_obj.tb_time = datetime.datetime.now()
                    if openid_obj.role:
                        pass
                    else:
                        openid_obj.role = 'xy'
                    # openid_obj.isActive = 'new'
                    # 2020.1.13 修改 关闭财务审核功能
                    openid_obj.isActive = 'true'

                    openid_obj.save()

                result = {'code': 200}
                return JsonResponse(result)
            else:
                result = {'code': 4013, 'error': 'Not enough permissions'}
                return JsonResponse(result)

    if request.method == 'PUT':

        openid_obj = Users.objects.get(openid=openid)

        if openid_obj.role == 'zs':
            # 小助手PUT改 课程

            # 判断前端是否给了json串
            json_str = request.body
            if not json_str:
                result = {'code': 241, 'error': 'Please give me data'}
                return JsonResponse(result)
            json_dic = json.loads(json_str)

            # 判断前端是否给了stu_id
            stu_id = json_dic.get('stu_id')
            if not stu_id:
                # 学生ID不能为空
                result = {'code': 242, 'error': 'Student id is None!'}
                return JsonResponse(result)

            # 判断学生是否存在
            try:
                stu_obj = Users.objects.get(id=stu_id)
            except:
                result = {'code': 243, 'error': 'Student in existence'}
                return JsonResponse(result)

            # 选择课程顾问
            if json_dic.get('counselor'):

                # 存
                stu_obj.counselor = json_dic.get('counselor')
                stu_obj.allotTime = json_dic.get('allotTime')
                stu_obj.save()

                result = {'code': 240}
                return JsonResponse(result)

            # 更改课程
            else:

                # 存
                if json_dic.get('feedbackTime'):
                    stu_obj.feedbackTime = json_dic.get('feedbackTime')
                    stu_obj.save()
                    del json_dic['feedbackTime']

                # if json_dic.get('note'):

                #     del json_dic['note']

                # 加
                del json_dic['openid']
                del json_dic['stu_id']
                for sta in json_dic:
                    # 遍历字典的key--> 状态
                    cous_id = json_dic[sta]
                    # 根据状态取出多个课程ID的列表 cous_id
                    for cou_id in cous_id:
                        # 遍历列表 取出课程ID cou_id

                        State.objects.create(
                            # 创建数据
                            state=sta,
                            course_id=cou_id,
                            users_id=stu_id,
                        )

                result = {'code': 240}
                return JsonResponse(result)

        if openid_obj.role == 'cw':
            json_str = request.body
            if not json_str:
                result = {'code': 231, 'error': 'Please give me data'}
                return JsonResponse(result)
            print(json_str)
            json_dic = json.loads(json_str)
            # print(json_dic)
            stu_id = json_dic.get('id')
            stu_obj = Users.objects.get(id=stu_id)
            judge = json_dic.get('judge')
            state = json_dic.get('state')
            if judge == 'register':
                # print(1)
                if state == 'true':
                    # print(2)
                    ambassador_name = json_dic.get('DSpicker')
                    if ambassador_name:
                        ambassador = Users.objects.get(username=ambassador_name)
                        stu_obj.ambassador = ambassador

                    if json_dic.get('BZpicker'):
                        if int(json_dic.get('BZpicker')) == 0:
                            stu_obj.currency = '人民币'
                        else:
                            stu_obj.currency = '加币'

                    if json_dic.get('count_wk'):
                        stu_obj.count_wk = json_dic.get('count_wk')

                    if json_dic.get('count_ms'):
                        stu_obj.count_ms = json_dic.get('count_ms')

                    if json_dic.get('Statepicker'):
                        payStates = json_dic.get('Statepicker')
                        if payStates == '全额':
                            stu_obj.payStates = 'qe'
                            stu_obj.pay_qe_time = json_dic.get('datechoice')
                        elif payStates == '申请费':
                            stu_obj.payStates = 'sq'
                            stu_obj.pay_sq_time = json_dic.get('datechoice')
                        elif payStates == '补齐':
                            stu_obj.payStates = 'bq'
                            stu_obj.pay_bq_time = json_dic.get('datechoice')

                    if json_dic.get('paid'):
                        if int(json_dic.get('BZpicker')) == 0:
                            paid = float(json_dic.get('paid')) / 5.2
                        else:
                            paid = float(json_dic.get('paid'))
                        stu_obj.paid = paid

                    if json_dic.get('balance'):
                        # print(json_dic.get('balance'))
                        if int(json_dic.get('BZpicker')) == 0:
                            balance = float(json_dic.get('balance')) / 5.2
                        else:
                            balance = float(json_dic.get('balance'))
                        stu_obj.balance = balance
                    stu_obj.isActive = 'true'
                    stu_obj.save()
                    result = {'code': 200}
                    return JsonResponse(result)

                elif state == 'false':
                    old_rea = stu_obj.reason
                    stu_obj.reason = old_rea + str(datetime.datetime.now().year) + '-' + str(
                        datetime.datetime.now().month) + '-' + str(datetime.datetime.now().day) + ' : ' + json_dic.get(
                        'reason') + '\n'
                    stu_obj.isActive = 'false'
                    stu_obj.save()
                    result = {'code': 200}
                    return JsonResponse(result)

                else:
                    result = {'code': 4014, 'error': 'Not enough permissions'}
                    return JsonResponse(result)

            elif judge == 'refund':
                if state == 'true':
                    stu_obj.isActive = 'false'
                    stu_obj.payStates = 'yt'
                    stu_obj.refund_time = datetime.datetime.now()
                    stu_obj.save()
                    result = {'code': 200}
                    return JsonResponse(result)

                elif state == 'false':
                    old_rea = stu_obj.reason
                    stu_obj.reason = old_rea + str(datetime.datetime.now().year) + '-' + str(
                        datetime.datetime.now().month) + '-' + str(datetime.datetime.now().day) + ' : ' + json_dic.get(
                        'reason') + '\n'
                    stu_obj.isActive = 'false'
                    stu_obj.save()
                    result = {'code': 200}
                    return JsonResponse(result)

                else:
                    result = {'code': 4015, 'error': 'Not enough permissions'}
                    return JsonResponse(result)

            elif judge == 'ds_reg':
                print(1)
                ambassador_name = json_dic.get('fzr')
                print(2)
                try:
                    print(3)
                    ambassador = Users.objects.get(username=ambassador_name)
                    stu_obj.ambassador = ambassador
                    print(4)
                except Exception as e:
                    return JsonResponse({'code': 404})

                role_dic = {
                    '大使': 'ds',
                    '市场负责人': 'scfz',
                    '全职区域负责人': 'qqfz',
                    '兼职区域负责人': 'jqfz',
                }
                if json_dic.get('zw') == '驳回':
                    stu_obj.isActive = ''
                else:
                    stu_obj.role = role_dic[json_dic.get('zw')]
                    if stu_obj.wechat:
                        stu_obj.isActive = 'true'
                    else:
                        stu_obj.isActive = ''

                stu_obj.save()
                result = {'code': 200}
                return JsonResponse(result)

            else:
                result = {'code': 4016, 'error': 'Not enough permissions'}
                return JsonResponse(result)

    if request.method == 'DELETE':

        openid_obj = Users.objects.get(openid=openid)

        if openid_obj.role == 'zs':
            # 小助手PUT改 课程

            # 判断前端是否给了json串
            json_str = request.body
            if not json_str:
                result = {'code': 241, 'error': 'Please give me data'}
                return JsonResponse(result)
            json_dic = json.loads(json_str)

            # 判断前端是否给了state_id
            state_id = json_dic.get('state_id')
            if not state_id:
                # state_id 不能为空
                result = {'code': 242, 'error': 'State id is None!'}
                return JsonResponse(result)

            # 判断state关系是否存在 存在则删除
            try:
                state_obj = State.objects.get(id=state_id)
                state_obj.delete()
            except:
                result = {'code': 243, 'error': 'Student in existence'}
                return JsonResponse(result)


def user_voucher(request, openid):
    '''
    凭证图片 添加 删除
    :param request:
    :param openid:
    :return:
    '''

    # POST提交
    # if request.method == 'GET':
    #     users = Users.objects.filter(openid=openid)
    #     if not users:
    #         result = {'code':208,'error':'The user is not existed!'}
    #         return JsonResponse(result)
    #
    #     list_path = os.listdir('./media/{}/voucher'.format(users[0].id))
    #     result = {'code': 200, 'id': users[0].id, 'path': list_path}
    #
    #     return JsonResponse(result)

    if request.method == 'POST':
        # print(openid)
        users = Users.objects.filter(openid=openid)
        # print(users)
        if not users:
            result = {'code': 208, 'error': 'The user is not existed!'}
            return JsonResponse(result)

        if request.FILES.get('voucher'):
            # 正常提交图片信息，进行存储

            users[0].voucher = request.FILES['voucher']
            users[0].save()
            list_path = os.listdir('./media/{}/voucher'.format(users[0].id))
            # print(list_path)
            result = {'code': 200, 'id': users[0].id, 'path': list_path}
            if users[0].isActive:
                # 2020.1.13 修改 关闭财务审核功能
                # users[0].isActive = 'new'
                users[0].isActive = 'true'
                users[0].save()

            return JsonResponse(result)
        else:
            # 没有提交图片信息
            result = {'code': 211, 'error': 'Please give me voucher'}
            return JsonResponse(result)

    if request.method == 'DELETE':
        users = Users.objects.filter(openid=openid)
        if not users:
            result = {'code': 208, 'error': 'The user is not existed!'}
            return JsonResponse(result)

        # 判断前端是否给了json串
        json_str = request.body
        if not json_str:
            result = {'code': 241, 'error': 'Please give me data'}
            return JsonResponse(result)
        json_dic = json.loads(json_str)

        # 判断前端是否给了path
        path = json_dic.get('path')

        if not path:
            # path 不能为空
            result = {'code': 242, 'error': 'path id is None!'}
            return JsonResponse(result)

        path = 'media/' + path
        os.remove(path)

        result = {'code': 200}
        return JsonResponse(result)

        # 判断state关系是否存在 存在则删除
        #


def card_img(request, openid):
    '''
    身份证图片 添加 删除
    :param request:
    :param openid:
    :return:
    '''

    # 上传图片思路
    # 1.前端-》 form提交 并且 content-type 要改成multipart/form-data
    # 2.后端只要拿到post提交，request.FILES['avatar']
    # 注意：由于目前django获取put请求的 multipart数据较为复杂，故改为post获取multipart数据

    if request.method == 'POST':
        users = Users.objects.filter(openid=openid)
        # print(users)
        if not users:
            result = {'code': 208, 'error': 'The user is not existed!'}
            return JsonResponse(result)

        if request.FILES.get('card_img'):
            # 正常提交图片信息，进行存储
            # print(1)

            try:
                # print(2)
                users[0].card_img = request.FILES['card_img']
                users[0].save()
                # print(3)
            except:
                pass
            list_path = os.listdir('./media/{}/card_img'.format(users[0].id))
            # print(list_path)
            result = {'code': 200, 'id': users[0].id, 'path': list_path}

            return JsonResponse(result)
        else:
            # 没有提交图片信息
            result = {'code': 211, 'error': 'Please give me card_img'}
            return JsonResponse(result)

    if request.method == 'DELETE':
        users = Users.objects.filter(openid=openid)
        if not users:
            result = {'code': 208, 'error': 'The user is not existed!'}
            return JsonResponse(result)

        # 判断前端是否给了json串
        json_str = request.body
        if not json_str:
            result = {'code': 241, 'error': 'Please give me data'}
            return JsonResponse(result)
        json_dic = json.loads(json_str)

        # 判断前端是否给了path
        path = json_dic.get('path')

        if not path:
            # path 不能为空
            result = {'code': 242, 'error': 'path id is None!'}
            return JsonResponse(result)

        path = 'media/' + path
        os.remove(path)

        result = {'code': 200}
        return JsonResponse(result)


def GPA_img(request, openid):
    '''
    LOP图片 添加 删除
    :param request:
    :param openid:
    :return:
    '''

    if request.method == 'POST':
        users = Users.objects.filter(openid=openid)
        # print(users)
        if not users:
            result = {'code': 208, 'error': 'The user is not existed!'}
            return JsonResponse(result)

        if request.FILES.get('GPA_img'):
            # 正常提交图片信息，进行存储
            # print(1)

            try:
                # print(2)
                users[0].GPA_img = request.FILES['GPA_img']
                users[0].save()
            except:
                pass
            list_path = os.listdir('./media/{}/GPA_img'.format(users[0].id))
            # print(list_path)
            result = {'code': 200, 'id': users[0].id, 'path': list_path}

            return JsonResponse(result)
        else:
            # 没有提交图片信息
            result = {'code': 211, 'error': 'Please give me GPA_img'}
            return JsonResponse(result)

    if request.method == 'DELETE':
        users = Users.objects.filter(openid=openid)
        if not users:
            result = {'code': 208, 'error': 'The user is not existed!'}
            return JsonResponse(result)

        # 判断前端是否给了json串
        json_str = request.body
        if not json_str:
            result = {'code': 241, 'error': 'Please give me data'}
            return JsonResponse(result)
        json_dic = json.loads(json_str)

        # 判断前端是否给了path
        path = json_dic.get('path')

        if not path:
            # path 不能为空
            result = {'code': 242, 'error': 'path id is None!'}
            return JsonResponse(result)

        path = 'media/' + path
        os.remove(path)

        result = {'code': 200}
        return JsonResponse(result)


def web_user_voucher(request, webopenid):
    '''
    凭证图片 添加 删除
    :param request:
    :param openid:
    :return:
    '''

    if request.method == 'POST':
        # print(webopenid)
        users = Users.objects.filter(webopenid=webopenid)
        # print(users[0].webopenid)
        if not users:
            result = {'code': 208, 'error': 'The user is not existed!'}
            return JsonResponse(result)

        # print(request.FILES)
        if request.FILES.get('file'):
            # 正常提交图片信息，进行存储

            users[0].voucher = request.FILES['file']
            users[0].save()
            list_path = os.listdir('./media/{}/voucher'.format(users[0].id))
            # print(list_path)
            result = {'code': 200, 'id': users[0].id, 'path': list_path}
            if users[0].isActive:
                # 2020.1.13 修改 关闭财务审核功能
                # users[0].isActive = 'new'
                users[0].isActive = 'true'
                users[0].save()

            return JsonResponse(result)
        else:
            # 没有提交图片信息
            result = {'code': 211, 'error': 'Please give me voucher'}
            return JsonResponse(result)

    if request.method == 'DELETE':
        users = Users.objects.filter(webopenid=webopenid)
        if not users:
            result = {'code': 208, 'error': 'The user is not existed!'}
            return JsonResponse(result)

        # 判断前端是否给了json串
        json_str = request.body
        if not json_str:
            result = {'code': 241, 'error': 'Please give me data'}
            return JsonResponse(result)
        json_dic = json.loads(json_str)

        # 判断前端是否给了path
        path = json_dic.get('path')

        if not path:
            # path 不能为空
            result = {'code': 242, 'error': 'path id is None!'}
            return JsonResponse(result)

        path = 'media/' + path
        os.remove(path)

        result = {'code': 200}
        return JsonResponse(result)

        # 判断state关系是否存在 存在则删除
        #


def web_card_img(request, webopenid):
    '''
    身份证图片 添加 删除
    :param request:
    :param openid:
    :return:
    '''

    # 上传图片思路
    # 1.前端-》 form提交 并且 content-type 要改成multipart/form-data
    # 2.后端只要拿到post提交，request.FILES['avatar']
    # 注意：由于目前django获取put请求的 multipart数据较为复杂，故改为post获取multipart数据

    if request.method == 'POST':
        users = Users.objects.filter(webopenid=webopenid)
        # print(users)
        if not users:
            result = {'code': 208, 'error': 'The user is not existed!'}
            return JsonResponse(result)

        if request.FILES.get('file'):
            # 正常提交图片信息，进行存储
            # print(1)

            try:
                # print(2)
                users[0].card_img = request.FILES['file']
                users[0].save()
                # print(3)
            except:
                pass
            list_path = os.listdir('./media/{}/card_img'.format(users[0].id))
            # print(list_path)
            result = {'code': 200, 'id': users[0].id, 'path': list_path}

            return JsonResponse(result)
        else:
            # 没有提交图片信息
            result = {'code': 211, 'error': 'Please give me card_img'}
            return JsonResponse(result)

    if request.method == 'DELETE':
        users = Users.objects.filter(webopenid=webopenid)
        if not users:
            result = {'code': 208, 'error': 'The user is not existed!'}
            return JsonResponse(result)

        # 判断前端是否给了json串
        json_str = request.body
        if not json_str:
            result = {'code': 241, 'error': 'Please give me data'}
            return JsonResponse(result)
        json_dic = json.loads(json_str)

        # 判断前端是否给了path
        path = json_dic.get('path')

        if not path:
            # path 不能为空
            result = {'code': 242, 'error': 'path id is None!'}
            return JsonResponse(result)

        path = 'media/' + path
        os.remove(path)

        result = {'code': 200}
        return JsonResponse(result)


def web_GPA_img(request, webopenid):
    '''
    LOP图片 添加 删除
    :param request:
    :param openid:
    :return:
    '''

    if request.method == 'POST':
        users = Users.objects.filter(webopenid=webopenid)
        # print(users)
        if not users:
            result = {'code': 208, 'error': 'The user is not existed!'}
            return JsonResponse(result)

        if request.FILES.get('file'):
            # 正常提交图片信息，进行存储
            # print(1)

            try:
                # print(2)
                users[0].GPA_img = request.FILES['file']
                users[0].save()
            except:
                pass
            list_path = os.listdir('./media/{}/GPA_img'.format(users[0].id))
            # print(list_path)
            result = {'code': 200, 'id': users[0].id, 'path': list_path}

            return JsonResponse(result)
        else:
            # 没有提交图片信息
            result = {'code': 211, 'error': 'Please give me GPA_img'}
            return JsonResponse(result)

    if request.method == 'DELETE':
        users = Users.objects.filter(webopenid=webopenid)
        if not users:
            result = {'code': 208, 'error': 'The user is not existed!'}
            return JsonResponse(result)

        # 判断前端是否给了json串
        json_str = request.body
        if not json_str:
            result = {'code': 241, 'error': 'Please give me data'}
            return JsonResponse(result)
        json_dic = json.loads(json_str)

        # 判断前端是否给了path
        path = json_dic.get('path')

        if not path:
            # path 不能为空
            result = {'code': 242, 'error': 'path id is None!'}
            return JsonResponse(result)

        path = 'media/' + path
        os.remove(path)

        result = {'code': 200}
        return JsonResponse(result)


def LOP_img(request, id):
    '''
    LOP图片 添加 删除
    :param request:
    :param openid:
    :return:
    '''
    if request.method == 'GET':
        users = Users.objects.filter(id=id)
        # print(users)
        if not users:
            result = {'code': 208, 'error': 'The user is not existed!'}
            return JsonResponse(result)
        try:
            list_path = os.listdir('./media/{}/LOP_img'.format(users[0].id))
        except:
            list_path = []
        result = {'code': 0, "msg": "", 'id': users[0].id, 'data': list_path}

        return JsonResponse(result)

    if request.method == 'POST':

        users = Users.objects.filter(id=id)
        # print(users)
        if not users:
            result = {'code': 208, 'error': 'The user is not existed!'}
            return JsonResponse(result)

        if request.FILES.get('LOP_img'):
            # 正常提交图片信息，进行存储
            # print(1)

            users[0].LOP_img = request.FILES['LOP_img']
            users[0].save()
            list_path = os.listdir('./media/{}/LOP_img'.format(users[0].id))
            # print(list_path)
            result = {'code': 0, "msg": "", 'id': users[0].id, 'data': list_path}

            return JsonResponse(result)
        else:
            # 没有提交图片信息
            result = {'code': 211, 'error': 'Please give me LOP_img'}
            return JsonResponse(result)

    if request.method == 'DELETE':
        users = Users.objects.filter(id=id)
        if not users:
            result = {'code': 208, 'error': 'The user is not existed!'}
            return JsonResponse(result)

        # 判断前端是否给了json串
        json_str = request.body
        if not json_str:
            result = {'code': 241, 'error': 'Please give me data'}
            return JsonResponse(result)
        json_dic = json.loads(json_str)

        # 判断前端是否给了path
        path = json_dic.get('path')

        if not path:
            # path 不能为空
            result = {'code': 242, 'error': 'path id is None!'}
            return JsonResponse(result)

        path = 'media/' + path
        os.remove(path)

        list_path = os.listdir('./media/{}/LOP_img'.format(users[0].id))
        # print(list_path)
        result = {'code': 0, "msg": "", 'id': users[0].id, 'data': list_path}

        return JsonResponse(result)


def GPA_imgs(request, id):
    '''
    GPA图片 添加 删除
    :param request:
    :param openid:
    :return:
    '''
    if request.method == 'GET':
        users = Users.objects.filter(id=id)
        # print(users)
        if not users:
            result = {'code': 208, 'error': 'The user is not existed!'}
            return JsonResponse(result)
        try:
            list_path = os.listdir('./media/{}/GPA_img'.format(users[0].id))
        except:
            list_path = []
        result = {'code': 0, "msg": "", 'id': users[0].id, 'data': list_path}

        return JsonResponse(result)

    if request.method == 'POST':

        users = Users.objects.filter(id=id)
        # print(users)
        if not users:
            result = {'code': 208, 'error': 'The user is not existed!'}
            return JsonResponse(result)

        if request.FILES.get('GPA_img'):
            # 正常提交图片信息，进行存储
            # print(1)

            users[0].GPA_img = request.FILES['GPA_img']
            users[0].save()
            list_path = os.listdir('./media/{}/GPA_img'.format(users[0].id))
            # print(list_path)
            result = {'code': 0, "msg": "", 'id': users[0].id, 'data': list_path}

            return JsonResponse(result)
        else:
            # 没有提交图片信息
            result = {'code': 211, 'error': 'Please give me GPA_img'}
            return JsonResponse(result)

    if request.method == 'DELETE':
        users = Users.objects.filter(id=id)
        if not users:
            result = {'code': 208, 'error': 'The user is not existed!'}
            return JsonResponse(result)

        # 判断前端是否给了json串
        json_str = request.body
        if not json_str:
            result = {'code': 241, 'error': 'Please give me data'}
            return JsonResponse(result)
        json_dic = json.loads(json_str)

        # 判断前端是否给了path
        path = json_dic.get('path')

        if not path:
            # path 不能为空
            result = {'code': 242, 'error': 'path id is None!'}
            return JsonResponse(result)

        path = 'media/' + path
        os.remove(path)

        list_path = os.listdir('./media/{}/GPA_img'.format(users[0].id))
        # print(list_path)
        result = {'code': 0, "msg": "", 'id': users[0].id, 'data': list_path}

        return JsonResponse(result)


def ds_registered(request, openid):
    if request.method == 'GET':
        openid_obj = Users.objects.get(openid=openid)
        if openid_obj.username:
            data = {
                'username': str(openid_obj.username),
                'gender': str(openid_obj.gender),
                'domesticTelephone': str(openid_obj.domesticTelephone),
                'foreignTelephone': str(openid_obj.foreignTelephone),
                'IDCard': str(openid_obj.IDCard)
            }
            result = {'code': 200, 'data': data}
            return JsonResponse(result)
        else:
            result = {'code': 207}
            return JsonResponse(result)

    if request.method == 'POST':
        openid_obj = Users.objects.get(openid=openid)
        if len(request.POST) > 1:
            openid_obj.username = request.POST.get('username')
            openid_obj.gender = request.POST.get('gender')
            openid_obj.domesticTelephone = request.POST.get('domesticTelephone')
            openid_obj.foreignTelephone = request.POST.get('foreignTelephone')
            openid_obj.IDCard = request.POST.get('IDCard')
            openid_obj.isActive = 'nds'
        else:
            openid_obj.isActive = 'nds'


# ---------------------------------------------- 小助手 ( 前后端不分离 ) --------------------------------------------------


def assistant_login(request):
    if request.method == 'GET':
        if request.session.get('userinfo'):
            ast_json = request.session.get('userinfo', '')
            ast_name = ast_json.get('ast_name', '')
            password = ast_json.get('password', '')
            try:
                ast = Assistant.objects.get(ast_name=ast_name, password=password)
                return render(request, 'login.html')
            except:
                return render(request, 'error.html')
        else:
            return render(request, 'login.html')


    elif request.method == 'POST':
        ast_name = request.POST.get('ast_name', '')
        password = request.POST.get('password', '')
        # 验证用户名，密码是否正确
        try:
            ast = Assistant.objects.get(ast_name=ast_name, password=password)
            # 在当前连接的Session中记录当前用户的信息
            request.session['userinfo'] = {
                "ast_name": ast.ast_name,
                "password": ast.password,
            }
            return render(request, 'login.html')

        except:
            return render(request, 'error.html')


def assistant(request):
    if request.method == 'GET':
        def time_formate(time):
            if time:
                return time
            else:
                return ''

        if request.GET.keys():
            orm = 'Users.objects.filter('
            last = ')'

            for key in request.GET.keys():

                if key == 'dp_search':
                    value = request.GET[key]
                    value_list = value.split(' ')
                    # print(value_list[1])
                    data = data_generate('Users', 'id=' + value_list[0], [value_list[1]])
                    result = {'code': 240, 'data': data}
                    return JsonResponse(result)

                if key == 'course_name':
                    value = request.GET[key]
                    data_list = []
                    for i in value.split():
                        data = field_data_generate('Course', 'course', 'id={}'.format(i))
                        data_list += data
                    result = {'code': 240, 'data': data_list}
                    return JsonResponse(result)

                if key == 'dp_menu':
                    value = request.GET[key]

                    if value == 'school_sk':
                        data = field_data_generate('School', 's_name', 'func="sk"')
                        # print(data)
                        result = {'code': 240, 'data': data}
                        return JsonResponse(result)

                    if value == 'school_sh':
                        data = field_data_generate('School', 's_name', 'func="sh"')
                        # print(data)
                        result = {'code': 240, 'data': data}
                        return JsonResponse(result)

                    if value == 'course':
                        data = field_data_generate('Course', 'course')
                        result = {'code': 240, 'data': data}
                        return JsonResponse(result)

                    if value == 'course_wk':
                        data = field_data_generate('Course', 'course', 'category="wk"')
                        result = {'code': 240, 'data': data}
                        return JsonResponse(result)

                    if value == 'course_ms':
                        data = field_data_generate('Course', 'course', 'category="ms"')
                        result = {'code': 240, 'data': data}
                        return JsonResponse(result)

                    if value == 'counselor':
                        data = field_data_generate('Counselor', 'counselor')
                        result = {'code': 240, 'data': data}
                        return JsonResponse(result)

                    # 其他字段均在Users表中
                    data = field_data_generate('Users', value)
                    result = {'code': 240, 'data': data}
                    return JsonResponse(result)

                if key == 'course_show':
                    try:
                        stu_obj = Users.objects.get(id=request.GET[key])
                        no_sub_wk = []
                        no_sub_ms = []
                        sub_wk = []
                        sub_ms = []
                        pass_wk = []
                        pass_ms = []
                        LOP_wk = []
                        LOP_ms = []

                        ns_course = State.objects.filter(state='no_sub', users=stu_obj)
                        s_course = State.objects.filter(state='sub', users=stu_obj)
                        p_course = State.objects.filter(state='pass', users=stu_obj)
                        L_course = State.objects.filter(state='LOP', users=stu_obj)

                        for i in ns_course:
                            if i.course.category == 'ms':
                                no_sub_ms.append([i.course.id, i.course.course])
                            if i.course.category == 'wk':
                                no_sub_wk.append([i.course.id, i.course.course])

                        for i in s_course:
                            if i.course.category == 'ms':
                                sub_ms.append([i.course.id, i.course.course])
                            if i.course.category == 'wk':
                                sub_wk.append([i.course.id, i.course.course])

                        for i in p_course:
                            if i.course.category == 'ms':
                                pass_ms.append([i.course.id, i.course.course])
                            if i.course.category == 'wk':
                                pass_wk.append([i.course.id, i.course.course])

                        for i in L_course:
                            if i.course.category == 'ms':
                                LOP_ms.append([i.course.id, i.course.course])
                            if i.course.category == 'wk':
                                LOP_wk.append([i.course.id, i.course.course])

                        # print(no_sub_ms)
                        # print(no_sub_wk)
                        # print(sub_ms)
                        # print(sub_wk)
                        # print(pass_ms)
                        # print(pass_wk)
                        # print(LOP_ms)
                        # print(LOP_wk)

                        data = {
                            'no_sub_ms': no_sub_ms,
                            'no_sub_wk': no_sub_wk,
                            'sub_ms': sub_ms,
                            'sub_wk': sub_wk,
                            'pass_ms': pass_ms,
                            'pass_wk': pass_wk,
                            'LOP_ms': LOP_ms,
                            'LOP_wk': LOP_wk,
                        }

                        result = {'code': 240, 'data': data}
                        return JsonResponse(result)

                        # print(stu_obj)
                    except:
                        result = {'code': 4017, 'error': 'Not enough permissions'}
                        return JsonResponse(result)

                if key == 'no_gw':
                    data = data_generate('Users', "isActive='true',counselor=None", ['id', 'username'])
                    result = {'code': 240, 'data': data}
                    return JsonResponse(result)

                if key == 'fuzzy_search':
                    # 根据查询关键词判断 从 邮箱 手机号 姓名中进行模糊查询
                    # 邮箱: 当查询关键词中有 @ 时 查询 firstEmail secondEmail 字段
                    # 手机号:  当查询关键词为纯数字时 查询 domesticTelephone foreignTelephone 字段
                    # 姓名:  其他情况 查询 username 字段
                    # data为满足条件对象的 姓名 邮箱 身份证 上课学校 网课 面授课

                    value = request.GET[key]
                    xzs_name = request.session.get('userinfo')['ast_name']
                    data = fuzzy_obj(value, xzs_name)
                    # print(data)
                    result = {'code': 240, 'data': data}
                    return JsonResponse(result)

                if key == 'stu_msg':
                    # 格式为 ?stu_msg=学号(必须放第一个) base(可有可无 必须放第二个 查询全量数据) 字段1 字段2 字段3 ..... ( 中间用空格隔开 )
                    # e.g: ?stu_msg=15 username id GPA

                    # 选择课程顾问 =学号 base 获取全量数据即可
                    # 已报名学生分配课程 =学号 base counselor allotTime course_ms/wk school_zs/sh/sk count_wk count_ms

                    value = request.GET[key]
                    value = value.split(' ')
                    s_id = value[0]
                    value = value[1:]
                    data = data_generate('Users', 'id={}'.format(s_id), value)
                    # print(data)
                    result = {'code': 240, 'data': data}
                    return JsonResponse(result)

                if key == 'all':
                    if request.GET[key] == 'my_all':
                        xzs_name = request.session.get('userinfo')['ast_name']
                        if xzs_name == "Paul":
                            data = Users.objects.filter(
                                Q(isActive='true') | Q(isActive='update') | Q(isActive='nds') | Q(isActive='refund'))

                        else:
                            data = Users.objects.filter(
                                Q(isActive='true') | Q(isActive='update') | Q(isActive='nds') | Q(isActive='refund'),
                                school__xzs=xzs_name)
                        data_list = []

                        for obj in data:

                            if obj.school.filter(func='sk'):
                                school_sk = obj.school.filter(func='sk')[0].s_name
                            else:
                                school_sk = ''

                            if obj.course_set.all():
                                course_wk_set = set()
                                course_ms_set = set()
                                for i in obj.course_set.filter(category='wk'):
                                    course_wk_set.add(i.course.split('--')[0])

                                for i in obj.course_set.filter(category='ms'):
                                    course_ms_set.add(i.course.split('--')[0])
                                    # print(i.course)

                                course_wk_list = list(course_wk_set)
                                course_ms_list = list(course_ms_set)

                            else:
                                course_wk_list = ''
                                course_ms_list = ''

                            course_wk_1 = ''
                            course_wk_2 = ''
                            course_wk_3 = ''
                            course_wk_4 = ''
                            course_wk_5 = ''
                            course_ms_1 = ''
                            course_ms_2 = ''
                            course_ms_3 = ''
                            course_ms_4 = ''
                            course_ms_5 = ''
                            courses_wk = [course_wk_1, course_wk_2, course_wk_3, course_wk_4, course_wk_5]
                            courses_ms = [course_ms_1, course_ms_2, course_ms_3, course_ms_4, course_ms_5]

                            for i in range(len(course_wk_list)):
                                courses_wk[i] = course_wk_list[i]

                            for i in range(len(course_ms_list)):
                                courses_ms[i] = course_ms_list[i]

                            # 大使
                            amb_user = obj.ambassador
                            if amb_user:
                                amb_user = amb_user.username
                            else:
                                amb_user = ''

                            # 缴费状态
                            pay_s_dic = {'sq': '仅缴申请费', 'bq': '已补齐缴费', 'qe': '全额缴费', 'yt': '已经退费', '': '未缴费'}
                            pay_s = pay_s_dic[str(obj.payStates)]

                            # school
                            if len(obj.school.filter(func='zs')) > 0:
                                zs_school = obj.school.filter(func='zs')[0].s_name
                            else:
                                zs_school = ''

                            if len(obj.school.filter(func='sh')) > 0:
                                sh_school = obj.school.filter(func='sh')[0].s_name
                            else:
                                sh_school = ''

                            # print(obj.allotTime)

                            data_list.append({'id': obj.stu_num,
                                              'stu_num': obj.stu_num,
                                              'username': '<a href="assistant_msg?search_msg={}">{}</a>'.format(
                                                  str(obj.id),
                                                  str(
                                                      obj.username)),
                                              'zs_school': zs_school,
                                              'sh_school': sh_school,
                                              'grade': obj.grade,
                                              'study': obj.study,
                                              'ambassador': amb_user,
                                              'firstEmail': obj.firstEmail, 'secondEmail': obj.secondEmail,
                                              'counselor': str(obj.counselor),
                                              'payStates': pay_s,
                                              'allotTime': time_formate(obj.allotTime),
                                              'subTime': time_formate(obj.subTime),
                                              'passTime': time_formate(obj.passTime),
                                              'LOPTime': time_formate(obj.LOPTime),
                                              'IDCard': obj.IDCard, 'sk_school': school_sk, 'wk_num': obj.count_wk,
                                              'ms_num': obj.count_ms, 'wechat': obj.wechat,
                                              'course_wk_1': courses_wk[0],
                                              'course_wk_2': courses_wk[1], 'course_wk_3': courses_wk[2],
                                              'course_wk_4': courses_wk[3],
                                              'course_wk_5': courses_wk[4], 'course_ms_1': courses_ms[0],
                                              'course_ms_2': courses_ms[1], 'course_ms_3': courses_ms[2],
                                              'course_ms_4': courses_ms[3], 'course_ms_5': courses_ms[4]})

                        # print(data_list)

                        result = {"code": 0, "msg": "", "count": 1000, "data": data_list}
                        return JsonResponse(result)
                    if request.GET[key] == 'all_all':
                        data = Users.objects.filter(
                            Q(isActive='true') | Q(isActive='update') | Q(isActive='nds') | Q(isActive='refund'),
                            ~Q(role='cw'))
                        data_list = []

                        for obj in data:

                            if obj.school.filter(func='sk'):
                                school_sk = obj.school.filter(func='sk')[0].s_name
                            else:
                                school_sk = ''

                            if obj.course_set.all():
                                course_wk_set = set()
                                course_ms_set = set()
                                for i in obj.course_set.filter(category='wk'):
                                    course_wk_set.add(i.course.split('--')[0])

                                for i in obj.course_set.filter(category='ms'):
                                    course_ms_set.add(i.course.split('--')[0])
                                    # print(i.course)

                                course_wk_list = list(course_wk_set)
                                course_ms_list = list(course_ms_set)

                            else:
                                course_wk_list = ''
                                course_ms_list = ''

                            course_wk_1 = ''
                            course_wk_2 = ''
                            course_wk_3 = ''
                            course_wk_4 = ''
                            course_wk_5 = ''
                            course_ms_1 = ''
                            course_ms_2 = ''
                            course_ms_3 = ''
                            course_ms_4 = ''
                            course_ms_5 = ''
                            courses_wk = [course_wk_1, course_wk_2, course_wk_3, course_wk_4, course_wk_5]
                            courses_ms = [course_ms_1, course_ms_2, course_ms_3, course_ms_4, course_ms_5]

                            for i in range(len(course_wk_list)):
                                courses_wk[i] = course_wk_list[i]

                            for i in range(len(course_ms_list)):
                                courses_ms[i] = course_ms_list[i]

                            # 大使
                            amb_user = obj.ambassador
                            if amb_user:
                                amb_user = amb_user.username
                            else:
                                amb_user = ''

                            # 缴费状态
                            pay_s_dic = {'sq': '仅缴申请费', 'bq': '已补齐缴费', 'qe': '全额缴费', 'yt': '已经退费', '': '未缴费'}
                            pay_s = pay_s_dic[str(obj.payStates)]

                            # school
                            if len(obj.school.filter(func='zs')) > 0:
                                zs_school = obj.school.filter(func='zs')[0].s_name
                            else:
                                zs_school = ''

                            if len(obj.school.filter(func='sh')) > 0:
                                sh_school = obj.school.filter(func='sh')[0].s_name
                            else:
                                sh_school = ''

                            # print(obj.allotTime)

                            data_list.append({'id': obj.stu_num,
                                              'stu_num': obj.stu_num,
                                              'username': '<a href="assistant_msg?search_msg={}">{}</a>'.format(
                                                  str(obj.id),
                                                  str(
                                                      obj.username)),
                                              'zs_school': zs_school,
                                              'sh_school': sh_school,
                                              'grade': obj.grade,
                                              'study': obj.study,
                                              'ambassador': amb_user,
                                              'firstEmail': obj.firstEmail, 'secondEmail': obj.secondEmail,
                                              'counselor': str(obj.counselor),
                                              'payStates': pay_s,
                                              'allotTime': time_formate(obj.allotTime),
                                              'subTime': time_formate(obj.subTime),
                                              'passTime': time_formate(obj.passTime),
                                              'LOPTime': time_formate(obj.LOPTime),
                                              'IDCard': obj.IDCard, 'sk_school': school_sk, 'wk_num': obj.count_wk,
                                              'ms_num': obj.count_ms, 'wechat': obj.wechat,
                                              'course_wk_1': courses_wk[0],
                                              'course_wk_2': courses_wk[1], 'course_wk_3': courses_wk[2],
                                              'course_wk_4': courses_wk[3],
                                              'course_wk_5': courses_wk[4], 'course_ms_1': courses_ms[0],
                                              'course_ms_2': courses_ms[1], 'course_ms_3': courses_ms[2],
                                              'course_ms_4': courses_ms[3], 'course_ms_5': courses_ms[4]})

                        # print(data_list)

                        result = {"code": 0, "msg": "", "count": 1000, "data": data_list}
                        return JsonResponse(result)
                    if request.GET[key] == 'qianfei':
                        xzs_name = request.session.get('userinfo')['ast_name']
                        if xzs_name == "Paul":
                            data = Users.objects.filter(
                                Q(isActive='true') | Q(isActive='update') | Q(isActive='nds') | Q(isActive='refund'),
                                ~Q(balance=0))
                        else:
                            data = Users.objects.filter(
                                Q(isActive='true') | Q(isActive='update') | Q(isActive='nds') | Q(isActive='refund'),
                                ~Q(balance=0), school__xzs=xzs_name)
                        data_list = []

                        for obj in data:

                            if obj.school.filter(func='sk'):
                                school_sk = obj.school.filter(func='sk')[0].s_name
                            else:
                                school_sk = ''

                            if obj.course_set.all():
                                course_wk_set = set()
                                course_ms_set = set()
                                for i in obj.course_set.filter(category='wk'):
                                    course_wk_set.add(i.course.split('--')[0])

                                for i in obj.course_set.filter(category='ms'):
                                    course_ms_set.add(i.course.split('--')[0])
                                    # print(i.course)

                                course_wk_list = list(course_wk_set)
                                course_ms_list = list(course_ms_set)

                            else:
                                course_wk_list = ''
                                course_ms_list = ''

                            course_wk_1 = ''
                            course_wk_2 = ''
                            course_wk_3 = ''
                            course_wk_4 = ''
                            course_wk_5 = ''
                            course_ms_1 = ''
                            course_ms_2 = ''
                            course_ms_3 = ''
                            course_ms_4 = ''
                            course_ms_5 = ''
                            courses_wk = [course_wk_1, course_wk_2, course_wk_3, course_wk_4, course_wk_5]
                            courses_ms = [course_ms_1, course_ms_2, course_ms_3, course_ms_4, course_ms_5]

                            for i in range(len(course_wk_list)):
                                courses_wk[i] = course_wk_list[i]

                            for i in range(len(course_ms_list)):
                                courses_ms[i] = course_ms_list[i]

                            # 大使
                            amb_user = obj.ambassador
                            if amb_user:
                                amb_user = amb_user.username
                            else:
                                amb_user = ''

                            # 缴费状态
                            pay_s_dic = {'sq': '仅缴申请费', 'bq': '已补齐缴费', 'qe': '全额缴费', 'yt': '已经退费', '': '未缴费'}
                            pay_s = pay_s_dic[str(obj.payStates)]

                            # school
                            if len(obj.school.filter(func='zs')) > 0:
                                zs_school = obj.school.filter(func='zs')[0].s_name
                            else:
                                zs_school = ''

                            if len(obj.school.filter(func='sh')) > 0:
                                sh_school = obj.school.filter(func='sh')[0].s_name
                            else:
                                sh_school = ''

                            # print(obj.allotTime)

                            data_list.append({'id': obj.stu_num,
                                              'stu_num': obj.stu_num,
                                              'username': '<a href="assistant_msg?search_msg={}">{}</a>'.format(
                                                  str(obj.id),
                                                  str(
                                                      obj.username)),
                                              'zs_school': zs_school,
                                              'sh_school': sh_school,
                                              'grade': obj.grade,
                                              'study': obj.study,
                                              'ambassador': amb_user,
                                              'firstEmail': obj.firstEmail, 'secondEmail': obj.secondEmail,
                                              'counselor': str(obj.counselor),
                                              'payStates': pay_s,
                                              'allotTime': time_formate(obj.allotTime),
                                              'subTime': time_formate(obj.subTime),
                                              'passTime': time_formate(obj.passTime),
                                              'LOPTime': time_formate(obj.LOPTime),
                                              'IDCard': obj.IDCard, 'sk_school': school_sk, 'wk_num': obj.count_wk,
                                              'ms_num': obj.count_ms, 'wechat': obj.wechat,
                                              'course_wk_1': courses_wk[0],
                                              'course_wk_2': courses_wk[1], 'course_wk_3': courses_wk[2],
                                              'course_wk_4': courses_wk[3],
                                              'course_wk_5': courses_wk[4], 'course_ms_1': courses_ms[0],
                                              'course_ms_2': courses_ms[1], 'course_ms_3': courses_ms[2],
                                              'course_ms_4': courses_ms[3], 'course_ms_5': courses_ms[4]})
                        # print(data_list)

                        result = {"code": 0, "msg": "", "count": 1000, "data": data_list}
                        return JsonResponse(result)

                if key == 'subTime':
                    time_list = request.GET[key].split(' - ')
                    # print("subTime__range=" + "('{}','{}')".format(time_list[0],time_list[1]))
                    orm += "subTime__range=" + '("{}","{}")'.format(time_list[0], time_list[1]) + ","


                elif key == 'passTime':
                    time_list = request.GET[key].split(' - ')
                    orm += "passTime__range=" + '("{}","{}")'.format(time_list[0], time_list[1]) + ","


                elif key == 'LOPTime':
                    time_list = request.GET[key].split(' - ')
                    orm += "LOPTime__range=" + '("{}","{}")'.format(time_list[0], time_list[1]) + ","


                elif key == 'school_sk':
                    if 'school_sk' in request.GET.keys() and 'school' in request.GET.keys():
                        last = ").filter(school='" + request.GET[key] + "')"
                    elif len(request.GET.keys()) == 3:
                        last = "(school='" + request.GET[key] + "')"
                    else:
                        last = ",school='" + request.GET[key] + "')"


                elif key == 'page':
                    pass

                elif key == 'limit':
                    pass

                else:
                    orm += key + "='" + request.GET[key] + "',"

            # print(orm)

            orm = orm[0:-1:1]
            orm += last
            # print(orm)

            try:
                data = eval(orm)

                # print(data)
                data = set(data)

                data_list = []

                for obj in data:

                    if obj.school.filter(func='sk'):
                        school_sk = obj.school.filter(func='sk')[0].s_name
                    else:
                        school_sk = ''

                    if obj.course_set.all():
                        course_wk_set = set()
                        course_ms_set = set()
                        for i in obj.course_set.filter(category='wk'):
                            course_wk_set.add(i.course.split('--')[0])

                        for i in obj.course_set.filter(category='ms'):
                            course_ms_set.add(i.course.split('--')[0])
                            # print(i.course)

                        course_wk_list = list(course_wk_set)
                        course_ms_list = list(course_ms_set)

                    else:
                        course_wk_list = ''
                        course_ms_list = ''

                    course_wk_1 = ''
                    course_wk_2 = ''
                    course_wk_3 = ''
                    course_wk_4 = ''
                    course_wk_5 = ''
                    course_ms_1 = ''
                    course_ms_2 = ''
                    course_ms_3 = ''
                    course_ms_4 = ''
                    course_ms_5 = ''
                    courses_wk = [course_wk_1, course_wk_2, course_wk_3, course_wk_4, course_wk_5]
                    courses_ms = [course_ms_1, course_ms_2, course_ms_3, course_ms_4, course_ms_5]

                    for i in range(len(course_wk_list)):
                        courses_wk[i] = course_wk_list[i]

                    for i in range(len(course_ms_list)):
                        courses_ms[i] = course_ms_list[i]

                    # 大使
                    amb_user = obj.ambassador
                    if amb_user:
                        amb_user = amb_user.username
                    else:
                        amb_user = ''

                    # 缴费状态
                    pay_s_dic = {'sq': '仅缴申请费', 'bq': '已补齐缴费', 'qe': '全额缴费', 'yt': '已经退费', '': '未缴费'}
                    pay_s = pay_s_dic[str(obj.payStates)]

                    # school
                    if len(obj.school.filter(func='zs')) > 0:
                        zs_school = obj.school.filter(func='zs')[0].s_name
                    else:
                        zs_school = ''

                    if len(obj.school.filter(func='sh')) > 0:
                        sh_school = obj.school.filter(func='sh')[0].s_name
                    else:
                        sh_school = ''

                    # print(obj.allotTime)

                    data_list.append({'id': obj.stu_num,
                                      'stu_num': obj.stu_num,
                                      'username': '<a href="assistant_msg?search_msg={}">{}</a>'.format(str(obj.id),
                                                                                                        str(
                                                                                                            obj.username)),
                                      'zs_school': zs_school,
                                      'sh_school': sh_school,
                                      'grade': obj.grade,
                                      'study': obj.study,
                                      'ambassador': amb_user,
                                      'firstEmail': obj.firstEmail, 'secondEmail': obj.secondEmail,
                                      'counselor': str(obj.counselor),
                                      'payStates': pay_s,
                                      'allotTime': time_formate(obj.allotTime),
                                      'subTime': time_formate(obj.subTime),
                                      'passTime': time_formate(obj.passTime),
                                      'LOPTime': time_formate(obj.LOPTime),
                                      'IDCard': obj.IDCard, 'sk_school': school_sk, 'wk_num': obj.count_wk,
                                      'ms_num': obj.count_ms, 'wechat': obj.wechat, 'course_wk_1': courses_wk[0],
                                      'course_wk_2': courses_wk[1], 'course_wk_3': courses_wk[2],
                                      'course_wk_4': courses_wk[3],
                                      'course_wk_5': courses_wk[4], 'course_ms_1': courses_ms[0],
                                      'course_ms_2': courses_ms[1], 'course_ms_3': courses_ms[2],
                                      'course_ms_4': courses_ms[3], 'course_ms_5': courses_ms[4]})

                result = {"code": 0, "msg": "", "count": 1000, "data": data_list}
                return JsonResponse(result)
                # return HttpResponse('OK')
            except Exception as e:
                print('------------------------------------', e)
                result = {'code': 4018, 'error': 'Not enough permissions'}
                return JsonResponse(result)

        else:
            xzs_name = request.session.get('userinfo')['ast_name']
            if xzs_name == "Paul":
                new_stu_num = Users.objects.filter(~Q(role='cw'), isActive='true', counselor=None)
            else:
                new_stu_num = Users.objects.filter(~Q(role='cw'), isActive='true', counselor=None, school__xzs=xzs_name)
            false_stu_num = Users.objects.filter(~Q(role='cw'), isActive='new')

            result = {'code': 240, 'data': {'new_stu_num': len(new_stu_num), 'false_stu_num': len(false_stu_num)}}
            return JsonResponse(result)

    if request.method == 'POST':
        try:
            # print(request.POST)
            id = request.POST['id']
            stu_obj = Users.objects.get(id=id)
            # print(stu_obj)
            # print(request.POST)
            for i in request.POST:
                # print(i+' 111111111111111111111111111')
                if hasattr(stu_obj, i) and request.POST.get(i, '') and i != 'counselor':

                    try:
                        setattr(stu_obj, i, request.POST[i])
                        stu_obj.save()
                    except:
                        print('保存出错')

                try:
                    if i == 'beizhu':
                        # print(i)
                        old_rea = stu_obj.reason
                        # print(str(datetime.datetime.now().year)+'-'+str(datetime.datetime.now().month)+'-'+str(datetime.datetime.now().day)+' : ' + old_rea + request.POST[i] + '\n')
                        stu_obj.reason = old_rea + str(datetime.datetime.now().year) + '-' + str(
                            datetime.datetime.now().month) + '-' + str(datetime.datetime.now().day) + ' : ' + \
                                         request.POST[i] + '\n'
                        stu_obj.save()
                except:
                    print('备注出错')

                try:
                    if i == 'stu_id':
                        stu_obj.stu_num = request.POST[i]
                        stu_obj.save()
                except:
                    print('学号出错')

                try:
                    if i == 'counselor':
                        # print(1)
                        counselor = Counselor.objects.get(id=request.POST[i])
                        # print(counselor)
                        stu_obj.counselor = counselor
                        stu_obj.save()
                except:
                    print('counselor')

                try:
                    if i == 'school_sh':
                        try:
                            school_obj = School.objects.get(id=request.POST[i])
                            school_obj.users_set.add(stu_obj)
                            # print(school_obj.users_set.all())
                            # school_obj._users = stu_obj
                        except:
                            pass
                except:
                    print('学校sh')

                try:
                    if i == 'school_sk':
                        try:
                            school_obj = School.objects.get(id=request.POST[i])
                            school_obj.users_set.add(stu_obj)
                            # print(school_obj.users_set.all())
                            # school_obj._users = stu_obj
                        except:
                            pass
                except:
                    print('学校sk')

                try:
                    if 'course' in i and request.POST.get(i, ''):
                        # print(i,request.POST[i])
                        if 'qx' in request.POST[i]:
                            # print(i, request.POST[i])
                            req_list = request.POST[i].split(' ')
                            cou_id = req_list[1]
                            # print(req_list)
                            delete = State.objects.filter(state='no_sub', users=stu_obj, course_id=cou_id)
                            delete.delete()
                        else:
                            exist = State.objects.filter(state='no_sub', users=stu_obj, course_id=request.POST[i])
                            # print(exist)
                            if not exist:
                                state = State()
                                state.state = 'no_sub'
                                state.users = stu_obj
                                state.course = Course.objects.get(id=request.POST[i])
                                state.save()
                except:
                    print('课程')

                try:
                    if 'sub' in i and request.POST.get(i, '') and i != 'subTime':
                        if 'qx' in request.POST[i]:
                            # print(i, request.POST[i])
                            req_list = request.POST[i].split(' ')
                            cou_id = req_list[1]
                            # print(req_list)
                            delete = State.objects.filter(state='sub', users=stu_obj, course_id=cou_id)
                            delete.delete()
                        else:
                            exist = State.objects.filter(state='sub', users=stu_obj, course_id=request.POST[i])
                            # print(exist)
                            if not exist:
                                state = State()
                                state.state = 'sub'
                                state.users = stu_obj
                                state.course = Course.objects.get(id=request.POST[i])
                                state.save()
                except:
                    print('提交')

                try:
                    if 'pass' in i and request.POST.get(i, '') and i != 'passTime':
                        if 'qx' in request.POST[i]:
                            # print(i, request.POST[i])
                            req_list = request.POST[i].split(' ')
                            cou_id = req_list[1]
                            # print(req_list)
                            delete = State.objects.filter(state='pass', users=stu_obj, course_id=cou_id)
                            delete.delete()
                        else:
                            exist = State.objects.filter(state='pass', users=stu_obj, course_id=request.POST[i])
                            # print(exist)
                            if not exist:
                                state = State()
                                state.state = 'pass'
                                state.users = stu_obj
                                state.course = Course.objects.get(id=request.POST[i])
                                state.save()
                except:
                    print('通过')

                try:
                    if 'LOP' in i and request.POST.get(i, '') and i != 'LOPTime':
                        if 'qx' in request.POST[i]:
                            # print(i, request.POST[i])
                            req_list = request.POST[i].split(' ')
                            cou_id = req_list[1]
                            # print(req_list)
                            delete = State.objects.filter(state='LOP', users=stu_obj, course_id=cou_id)
                            delete.delete()
                        else:
                            exist = State.objects.filter(state='LOP', users=stu_obj, course_id=request.POST[i])
                            # print(exist)
                            if not exist:
                                state = State()
                                state.state = 'LOP'
                                state.users = stu_obj
                                state.course = Course.objects.get(id=request.POST[i])
                                state.save()
                except:
                    print('LOP')

                # print(i)

            # return redirect('assistant_msg?search_msg={}'.format(id))
            return redirect('/v1/assistant_msg?search_msg={}'.format(id))
        except Exception as e:
            print(e)
            return render(request, 'error.html')

    # 无查询字符串返回 新学生(isActive为true counselor为NULL) 的个数


def assistant_index(request):
    '''
    首页
    :param request:
    :return:
    '''

    if request.session.get('userinfo'):
        ast_json = request.session.get('userinfo', '')
        ast_name = ast_json.get('ast_name', '')
        password = ast_json.get('password', '')
        try:
            ast = Assistant.objects.get(ast_name=ast_name, password=password)
            return render(request, 'assistant_index.html')
        except:
            return render(request, 'error.html')
    else:
        return render(request, 'error.html')


def assistant_search(request):
    '''
    已报名学生 搜索页
    :param request:
    :return:
    '''
    if request.GET.keys():
        # print(request.GET)
        for key in request.GET.keys():
            if key == 'search_msg':
                value = request.GET[key]
                print(1111111111111111111111111111111111111111111111111111)
                data = fuzzy_query_generate(value)
                result = {'code': 240, 'data': data}
                return JsonResponse(result)
    else:
        if request.session.get('userinfo'):
            ast_json = request.session.get('userinfo', '')
            ast_name = ast_json.get('ast_name', '')
            password = ast_json.get('password', '')
            try:
                ast = Assistant.objects.get(ast_name=ast_name, password=password)
                return render(request, 'stu_search.html')
            except:
                return render(request, 'error.html')
        else:
            return render(request, 'error.html')


def assistant_school(request):
    '''
    校区 搜索页
    :param request:
    :return:
    '''

    if request.session.get('userinfo'):
        ast_json = request.session.get('userinfo', '')
        ast_name = ast_json.get('ast_name', '')
        password = ast_json.get('password', '')
        try:
            ast = Assistant.objects.get(ast_name=ast_name, password=password)
            return render(request, 'sch_search.html')
        except:
            return render(request, 'error.html')
    else:
        return render(request, 'error.html')


def assistant_new(request):
    '''
    新学生页
    :param request:
    :return:
    '''

    if request.session.get('userinfo'):
        ast_json = request.session.get('userinfo', '')
        ast_name = ast_json.get('ast_name', '')
        password = ast_json.get('password', '')
        try:
            xzs_name = request.session.get('userinfo')['ast_name']
            # 验证
            ast = Assistant.objects.get(ast_name=ast_name, password=password)
            if xzs_name == "Paul":
                new_stu = Users.objects.filter(~Q(role='cw'), isActive='true', counselor=None)
            else:
                new_stu = Users.objects.filter(~Q(role='cw'), isActive='true', counselor=None, school__xzs=xzs_name)
            data = []
            for i in new_stu:
                school = data_generate('Users', 'id={}'.format(i.id), ['school_zs'])
                # print(school['school_zs'])
                data.append({'first': '报名序号 : ' + str(i.id), 'last': '姓名 : ' + str(i.username),
                             'school': '学校 : ' + school['school_zs']})
            return render(request, 'new_stu.html', {'data': data, 'title': '未分配顾问学生'})
        except:
            return render(request, 'error.html')
    else:
        return render(request, 'error.html')


def assistant_false(request):
    if request.session.get('userinfo'):
        ast_json = request.session.get('userinfo', '')
        ast_name = ast_json.get('ast_name', '')
        password = ast_json.get('password', '')
        try:
            xzs_name = request.session.get('userinfo')['ast_name']
            # 验证
            ast = Assistant.objects.get(ast_name=ast_name, password=password)
            false_stu = Users.objects.filter(~Q(role='cw'), isActive='new').order_by('tb_time')
            data = []
            for i in false_stu:
                school = data_generate('Users', 'id={}'.format(i.id), ['school_zs'])
                # print(school['school_zs'])
                data.append({'first': '报名序号 : ' + str(i.id), 'last': '姓名 : ' + str(i.username),
                             'school': '学校 : ' + school['school_zs']})
            return render(request, 'new_stu.html', {'data': data, 'title': '未通过学生'})
        except:
            return render(request, 'error.html')
    else:
        return render(request, 'error.html')


def assistant_distribution(request):
    '''
    学生分配顾问页
    :param request:
    :return:
    '''

    if request.session.get('userinfo'):
        ast_json = request.session.get('userinfo', '')
        ast_name = ast_json.get('ast_name', '')
        password = ast_json.get('password', '')
        try:
            ast = Assistant.objects.get(ast_name=ast_name, password=password)
            if request.GET.keys():
                for key in request.GET.keys():
                    if key == 'id':
                        value = request.GET[key]
                        value2 = value.strip().split(' : ')[1] + ' base school_zs'
                        stu_id = value.strip().split(' : ')[1]
                        try:
                            list_path = os.listdir('./media/{}/GPA_img'.format(stu_id))
                        except:
                            list_path = []

                        list_dic = {}

                        for i in range(len(list_path)):
                            list_dic['a' + str(i)] = list_path[i]
                        # print(value)
                        # return render(request, 'stu_distribution.html', {'data': json.dumps(value2)})
                        return render(request, 'stu_distribution.html',
                                      {'data': json.dumps(value2), 'stu_id': stu_id, 'list_path': list_dic})

        except:
            return render(request, 'error.html')
    else:
        return render(request, 'error.html')


def assistant_msg(request):
    '''
    学生详细信息页
    :param request:
    :return:
    '''

    if request.session.get('userinfo'):
        ast_json = request.session.get('userinfo', '')
        ast_name = ast_json.get('ast_name', '')
        password = ast_json.get('password', '')
        try:
            ast = Assistant.objects.get(ast_name=ast_name, password=password)
            if request.GET.keys():
                for key in request.GET.keys():
                    if key == 'search_msg':
                        value = request.GET[key]
                        try:
                            list_path = os.listdir('./media/{}/GPA_img'.format(value))
                        except:
                            list_path = []

                        list_dic = {}

                        for i in range(len(list_path)):
                            list_dic['a' + str(i)] = list_path[i]

                        return render(request, 'stu_msg.html', {'data': value, 'list_path': list_dic})
        except:
            return render(request, 'error.html')
    else:
        return render(request, 'error.html')


def assistant_sch_msg(request):
    if request.session.get('userinfo'):
        ast_json = request.session.get('userinfo', '')
        ast_name = ast_json.get('ast_name', '')
        password = ast_json.get('password', '')
        try:
            search_msg = request.GET['search_msg']
            # print(search_msg)
            return render(request, 'sch_msg.html', {'data': search_msg})
        except:
            return render(request, 'error.html')
    else:
        return render(request, 'error.html')


def test(request):
    school_list = ['阿尔伯塔大学', '阿萨巴斯卡大学', '卡尔加里大学', '莱斯桥大学', '加拿大皇家大学', '西蒙佛蕾泽大学', '不列颠哥伦比亚大学', '北哥伦比亚大学', '维多利亚大学',
                   '布兰顿大学',
                   '曼尼托巴大学', '温尼伯大学', '艾里森山大学', '圣托玛斯大学', '曼克顿大学', '新布朗斯维克大学', '纽芬兰纪念大学', '阿卡迪亚大学', '戴尔豪斯大学', '圣温森特山大学',
                   '圣安娜大学',
                   '圣法兰西斯-萨维尔大学', '圣玛莉大学', '布兰顿大学学院', '布鲁克大学', '卡尔顿大学', '湖岸大学', '劳伦特大学', '麦克马斯特大学', '尼波星大学', '女王大学',
                   '怀雅逊大学',
                   '特莱特大学', '圭尔夫大学', '渥太华大学', '多伦多大学', '滑铁卢大学', '西安大略大学', '温莎大学', '韦尔福瑞德-劳瑞大学', '约克大学', '爱德华王子岛大学',
                   '主教大学', '协和大学',
                   '拉瓦尔大学', '麦吉尔大学', '蒙特利尔大学', '魁北克大学', '魁北克大学校本部', '魁北克大学蒙特利尔分校', '魁北克大学特瓦河分校', '魁北克大学史库蒂密分校',
                   '魁北克大学瑞姆斯基分校',
                   '施尔布鲁克大学', '里贾纳大学', '萨斯卡彻温大学']

    # for i in school_list:
    #         print('已创建', i)
    #         # market = Market.objects.get(id=3)
    #         add = School()
    #         add.s_name = i
    #         add.func = 'zs'
    #         # add.market = market
    #         add.save()

    return render(request, 'test.html')


# ---------------------------------------------- 财务 ( 前后端不分离 ) --------------------------------------------------
def cw_login(request):
    if request.method == 'GET':
        # print(0)
        if request.session.get('userinfo'):
            ast_json = request.session.get('userinfo', '')
            ast_name = ast_json.get('ast_name', '')
            password = ast_json.get('password', '')
            try:
                ast = Assistant.objects.get(ast_name=ast_name, password=password)
                return render(request, 'cw_login.html')
            except:
                return render(request, 'cw_error.html')
        else:
            return render(request, 'cw_login.html')
    if request.method == 'POST':
        # print(1)
        ast_name = request.POST.get('ast_name', '')
        password = request.POST.get('password', '')
        # 验证用户名，密码是否正确
        try:
            # print(2)
            ast = Assistant.objects.get(ast_name=ast_name, password=password)
            # 在当前连接的Session中记录当前用户的信息
            request.session['userinfo'] = {
                "ast_name": ast.ast_name,
                "password": ast.password,
            }
            # print(3)
            return render(request, 'cw_login.html')

        except:
            return render(request, 'cw_error.html')


def cw_index(request):
    if request.session.get('userinfo'):
        ast_json = request.session.get('userinfo', '')
        ast_name = ast_json.get('ast_name', '')
        password = ast_json.get('password', '')
        try:
            # ast = Assistant.objects.get(ast_name=ast_name, password=password)
            return render(request, 'cw_index.html', {'ast_name': ast_name})
        except:
            return render(request, 'cw_error.html')
    else:
        return render(request, 'cw_error.html')


def cw_search(request):
    if request.session.get('userinfo'):
        ast_json = request.session.get('userinfo', '')
        ast_name = ast_json.get('ast_name', '')
        password = ast_json.get('password', '')
        try:
            # ast = Assistant.objects.get(ast_name=ast_name, password=password)
            return render(request, 'cw_ss_msg.html', {'ast_name': ast_name, 'title': '学生'})
        except:
            return render(request, 'cw_error.html')
    else:
        return render(request, 'cw_error.html')

    # return render(request,'cw_ss_msg.html',{'ast_name':'testts','title':'已报名学生'})


def cw_ds(request):
    if request.session.get('userinfo'):
        ast_json = request.session.get('userinfo', '')
        ast_name = ast_json.get('ast_name', '')
        password = ast_json.get('password', '')
        try:
            # ast = Assistant.objects.get(ast_name=ast_name, password=password)
            return render(request, 'cw_ss_msg.html', {'ast_name': ast_name, 'title': '大使'})
        except:
            return render(request, 'cw_error.html')
    else:
        return render(request, 'cw_error.html')

    # return render(request,'cw_ss_msg.html',{'ast_name':'testts','title':'大使列表'})


def cw_new(request):
    if request.session.get('userinfo'):
        ast_json = request.session.get('userinfo', '')
        ast_name = ast_json.get('ast_name', '')
        password = ast_json.get('password', '')
        try:
            # ast = Assistant.objects.get(ast_name=ast_name, password=password)
            return render(request, 'cw_jb_msg.html', {'ast_name': ast_name, 'title': '新学生'})
        except:
            return render(request, 'cw_error.html')
    else:
        return render(request, 'cw_error.html')

    # return render(request,'cw_jb_msg.html',{'ast_name':'testts','title':'新学生'})


def cw_nds(request):
    if request.session.get('userinfo'):
        ast_json = request.session.get('userinfo', '')
        ast_name = ast_json.get('ast_name', '')
        password = ast_json.get('password', '')
        try:
            # ast = Assistant.objects.get(ast_name=ast_name, password=password)
            return render(request, 'cw_jb_msg.html', {'ast_name': ast_name, 'title': '新大使申请'})
        except:
            return render(request, 'cw_error.html')
    else:
        return render(request, 'cw_error.html')

    # return render(request,'cw_jb_msg.html',{'ast_name':'testts','title':'新大使申请'})


def cw_refunds(request):
    if request.session.get('userinfo'):
        ast_json = request.session.get('userinfo', '')
        ast_name = ast_json.get('ast_name', '')
        password = ast_json.get('password', '')
        try:
            # ast = Assistant.objects.get(ast_name=ast_name, password=password)
            return render(request, 'cw_jb_msg.html', {'ast_name': ast_name, 'title': '退费申请'})
        except:
            return render(request, 'cw_error.html')
    else:
        return render(request, 'cw_error.html')

    # return render(request,'cw_jb_msg.html',{'ast_name':'testts','title':'退费申请'})


def cw_new_msg(request, stu_id=None):
    try:
        list_path = os.listdir('./media/{}/voucher'.format(stu_id))
    except:
        list_path = []

    list_dic = {}

    for i in range(len(list_path)):
        list_dic['a' + str(i)] = list_path[i]

    if request.session.get('userinfo'):
        ast_json = request.session.get('userinfo', '')
        ast_name = ast_json.get('ast_name', '')
        password = ast_json.get('password', '')
        try:
            # ast = Assistant.objects.get(ast_name=ast_name, password=password)
            return render(request, 'cw_new_msg.html', {'stu_id': stu_id, 'ast_name': ast_name, 'list_path': list_dic})
        except:
            return render(request, 'cw_error.html')
    else:
        return render(request, 'cw_error.html')

    # print(list_dic)

    # return render(request,'cw_new_msg.html',{'stu_id':stu_id,'ast_name':'testts','list_path':list_dic})


def cw_stu_msg(request, stu_id=None):
    try:
        list_path = os.listdir('./media/{}/voucher'.format(stu_id))
    except:
        list_path = []

    list_dic = {}

    for i in range(len(list_path)):
        list_dic['a' + str(i)] = list_path[i]

    if request.session.get('userinfo'):
        ast_json = request.session.get('userinfo', '')
        ast_name = ast_json.get('ast_name', '')
        password = ast_json.get('password', '')
        try:
            # ast = Assistant.objects.get(ast_name=ast_name, password=password)
            return render(request, 'cw_stu_msg.html', {'stu_id': stu_id, 'ast_name': ast_name, 'list_path': list_dic})
        except:
            return render(request, 'cw_error.html')
    else:
        return render(request, 'cw_error.html')

    # return render(request,'cw_stu_msg.html',{'stu_id':stu_id,'ast_name':'testts'})


def cw_ds_msg(request, stu_id=None):
    if request.session.get('userinfo'):
        ast_json = request.session.get('userinfo', '')
        ast_name = ast_json.get('ast_name', '')
        password = ast_json.get('password', '')
        try:
            # ast = Assistant.objects.get(ast_name=ast_name, password=password)
            return render(request, 'cw_ds_msg.html', {'stu_id': stu_id, 'ast_name': ast_name})
        except:
            return render(request, 'cw_error.html')
    else:
        return render(request, 'cw_error.html')

    # return render(request,'cw_ds_msg.html',{'stu_id':stu_id,'ast_name':'testts'})


def cw_new_ds(request, stu_id=None):
    if request.session.get('userinfo'):
        ast_json = request.session.get('userinfo', '')
        ast_name = ast_json.get('ast_name', '')
        password = ast_json.get('password', '')
        try:
            # ast = Assistant.objects.get(ast_name=ast_name, password=password)
            return render(request, 'cw_new_ds.html', {'stu_id': stu_id, 'ast_name': ast_name})
        except:
            return render(request, 'cw_error.html')
    else:
        return render(request, 'cw_error.html')

    # return render(request,'cw_new_ds.html',{'stu_id':stu_id,'ast_name':'testts'})


def cw_ref_msg(request, stu_id=None):
    try:
        list_path = os.listdir('./media/{}/voucher'.format(stu_id))
    except:
        list_path = []

    list_dic = {}

    for i in range(len(list_path)):
        list_dic['a' + str(i)] = list_path[i]

    if request.session.get('userinfo'):
        ast_json = request.session.get('userinfo', '')
        ast_name = ast_json.get('ast_name', '')
        password = ast_json.get('password', '')
        try:
            # ast = Assistant.objects.get(ast_name=ast_name, password=password)
            return render(request, 'cw_ref_msg.html', {'stu_id': stu_id, 'ast_name': ast_name, 'list_path': list_dic})
        except:
            return render(request, 'cw_error.html')
    else:
        return render(request, 'cw_error.html')

    # return render(request,'cw_ref_msg.html',{'stu_id':stu_id,'ast_name':'testts','list_path':list_dic})


def map_test(request):
    return render(request, 'map_test.html')


def c_v_script(request):

    # try:
    #     headers = {
    #         'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; WOW64; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; InfoPath.3)',
    #         'Cookie': 'Hm_lvt_f2f9a68d9323548d83786738cb6a2aeb=1574752405; MoodleSession=eab27d24bs8cpg3vj0ls5k6jp1'}
    #     cookie = {'Hm_lvt_f2f9a68d9323548d83786738cb6a2aeb': '1574752405', ' MoodleSession': 'eab27d24bs8cpg3vj0ls5k6jp1'}
    #
    #     xueke_url = 'http://moodle.canadasummer.ca/course/index.php?categoryid=1'
    #     xueke = requests.get(url=xueke_url, headers=headers, cookies=cookie)
    #     xueke_html = etree.HTML(xueke.text)
    #     a_list = xueke_html.xpath('//h3[@class="categoryname"]/a/@href')
    #     # a_text = xueke_html.xpath('//h3[@class="categoryname"]/a/text()')
    #
    #     for i in range(len(a_list)):
    #
    #         # print('学科名：',a_text[i])
    #         kecheng_url = a_list[i]
    #         time.sleep(random.uniform(0, 0.5))
    #         kecheng = requests.get(url=kecheng_url, headers=headers, cookies=cookie)
    #         kecheng_html = etree.HTML(kecheng.text)
    #         b_list = kecheng_html.xpath('//h3[@class="coursename"]/a/@href')
    #         b_text = kecheng_html.xpath('//h3[@class="coursename"]/a/text()')
    #         for i2 in range(len(b_list)):
    #
    #             # print('课程名',b_text[i2])
    #
    #             courser_obj = MoodleCourse.objects.filter(c_name__icontains=b_text[i2])
    #
    #             if len(courser_obj) == 1:
    #
    #                 topic_url = b_list[i2]
    #                 time.sleep(random.uniform(0, 0.5))
    #                 topic = requests.get(url=topic_url, headers=headers, cookies=cookie)
    #                 topic_html = etree.HTML(topic.text)
    #                 c_list = topic_html.xpath(
    #                     '//img[@src="http://moodle.canadasummer.ca/theme/image.php/_s/bcu/core/1567408182/f/mpeg-24"]/../@href')
    #                 # print(c_list)
    #
    #                 for i3 in range(len(c_list)):
    #                     video_url = c_list[i3]
    #                     # print(video_url)
    #
    #                     time.sleep(random.uniform(0, 0.5))
    #                     video = requests.get(url=video_url, headers=headers, cookies=cookie)
    #                     # print(video)
    #                     video_html = etree.HTML(video.text)
    #                     video_oss = video_html.xpath('//video/source/@src')
    #                     video_name = video_html.xpath('//span[@id="maincontent"]/../h2/text()')
    #
    #
    #                     # print('学科：', a_text[i], '课程', b_text[i2], '视频名称', video_name[0], '视频url', video_oss[0])
    #
    #                     try:
    #                         v_name = video_name[0]
    #                     except:
    #                         v_name = ''
    #
    #                     try:
    #                         v_oss = video_oss[0]
    #                     except:
    #                         v_oss = ''
    #
    #                     new_video = MoodleVideo()
    #                     new_video.v_name = v_name
    #                     new_video.video_add = v_oss
    #                     new_video.moodle_course = courser_obj[0]
    #
    #                     new_video.save()
    #
    #
    #     return HttpResponse('OK')
    #
    # except Exception as e :
    #     return HttpResponse(e)

    if request.method == 'GET':
        return render(request,'moodle/excel_video.html')

    if request.method == 'POST':

        course_list = []

        try:

            book = openpyxl.load_workbook(request.FILES['excel'])
            sheet = book.active

            # for row in sheet:
            #     # print(i[0].value, i[1].value, i[2].value)
            # 
            #     # m_course = MoodleCourse.objects.filter(c_name__icontains=i[0].value)
            #     # new_video = MoodleVideo()
            #     # new_video.v_name = i[1].value
            #     # new_video.video_add = i[2].value
            #     # new_video.moodle_course = m_course[0]
            #     #
            #     # new_video.save()
            # 
            #     if row[0].value in subject_list:
            # 
            #         if row[1].value in course_list:
            # 
            #             m_course = MoodleCourse.objects.filter(c_name__icontains=i[0].value)
            #             new_video = MoodleVideo()
            #             new_video.v_name = i[1].value
            #             new_video.video_add = i[2].value
            #             new_video.moodle_course = m_course[0]
            # 
            #             new_video.save()

            row = list(sheet.iter_rows())
            # col = list(sheet.iter_cols())
            #
            # for cou in col[1]:
            #     if cou.value not in course_list:
            #
            #         new_course = MoodleCourse()
            #
            #
            #         course_list.append(cou.value)




            for r in row:
                e_sub = r[0].value
                e_course = r[1].value
                e_name = r[2].value
                e_path = r[3].value

                if e_course not in course_list:
                    m_sub = MoodleSubject.objects.filter(s_name__icontains=e_sub)[0]

                    new_course = MoodleCourse()
                    new_course.c_name = e_course
                    new_course.subject = m_sub

                    new_course.save()

                    course_list.append(e_course)

                m_course = MoodleCourse.objects.filter(c_name__icontains=e_course)[0]

                new_video = MoodleVideo()
                new_video.v_name = e_name
                new_video.video_add = e_path
                new_video.moodle_course = m_course

                new_video.save()




            return HttpResponse('OK')

        except Exception as e:

            return HttpResponse(e)